import os
import uuid
import json
import re
import sys
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote


REPO_ROOT = Path(__file__).resolve().parents[2]
SERVICE_ROOT = REPO_ROOT / "services" / "waybill-parser"
if str(SERVICE_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVICE_ROOT))

TEST_DB_DIR = REPO_ROOT / ".pytest_cache"
TEST_DB_DIR.mkdir(exist_ok=True)
TEST_DB = TEST_DB_DIR / f"product_matching_stage82b_test_{uuid.uuid4().hex}.db"

os.environ["DATABASE_URL"] = f"sqlite:///{TEST_DB.as_posix()}"
os.environ["AUTO_CREATE_TABLES"] = "true"
os.environ["SECRET_KEY"] = "product-matching-stage82b-secret"

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from app.main import app  # noqa: E402
from app.models import (  # noqa: E402
    CaptureTask,
    ImageAsset,
    Product,
    ProductMatchingRule,
    ProductSku,
    RawCaptureRecord,
    RecognitionRulePack,
    StandardDetail,
    StandardDetailBatch,
)
from app.services import order_row_reader as order_row_reader_service  # noqa: E402
from service_app.order_row_engine import (  # noqa: E402
    draft_rows_from_standard_detail_values,
    draft_rows_from_waybill_sample,
    order_row_draft_summary,
)


def _headers(client: TestClient) -> dict[str, str]:
    login = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
    assert login.status_code == 200
    token = login.json()["access_token"]
    return {"Authorization": f"Bearer {token}", "X-Workspace-Id": "1"}


def _use_parser_service_stub(monkeypatch) -> None:
    def fake_parse_order_row_drafts_with_service(
        *,
        task_id: int,
        standard_details: list[dict] | None = None,
        raw_records: list[dict] | None = None,
        waybill_samples: list[dict] | None = None,
        rule_pack: dict,
    ) -> dict:
        assert raw_records in (None, [])
        assert rule_pack["pack"]["code"] == "test-shoe-pack"
        parents = [
            draft_rows_from_standard_detail_values(
                parser_input["field_values"],
                standard_detail_id=parser_input["standard_detail_id"],
                parent_sequence=parser_input["parent_sequence"],
            )
            for parser_input in (standard_details or [])
        ]
        parents.extend(
            draft_rows_from_waybill_sample(
                parser_input,
                parent_sequence=parser_input["parent_sequence"],
            )
            for parser_input in (waybill_samples or [])
        )
        return {
            "contract_version": "order_row_drafts_v1",
            "task_id": task_id,
            "summary": order_row_draft_summary(parents),
            "parents": [parent.as_dict() for parent in parents],
            "rows": [row.as_dict() for parent in parents for row in parent.rows],
        }

    monkeypatch.setattr(order_row_reader_service, "waybill_parser_service_enabled", lambda: True, raising=False)
    monkeypatch.setattr(
        order_row_reader_service,
        "parse_order_row_drafts_with_service",
        fake_parse_order_row_drafts_with_service,
        raising=False,
    )


def _seed_assets_and_detail() -> tuple[int, int, int, int]:
    with SessionLocal() as db:
        product = db.query(Product).filter(
            Product.workspace_id == 1,
            Product.name == "帆布鞋",
            Product.is_deleted.is_(False),
        ).first()
        if product is None:
            product = Product(
                tenant_id=1,
                workspace_id=1,
                name="帆布鞋",
                code="shoe",
                keywords=["鞋"],
                is_enabled=True,
            )
            db.add(product)
            db.flush()

        image = db.query(ImageAsset).filter(
            ImageAsset.workspace_id == 1,
            ImageAsset.name == "黑色图",
            ImageAsset.is_deleted.is_(False),
        ).first()
        if image is None:
            image = ImageAsset(
                tenant_id=1,
                workspace_id=1,
                name="黑色图",
                file_path="storage/black.png",
            )
            db.add(image)
            db.flush()

        sku = db.query(ProductSku).filter(
            ProductSku.workspace_id == 1,
            ProductSku.product_id == product.id,
            ProductSku.name == "黑色",
            ProductSku.is_deleted.is_(False),
        ).first()
        if sku is None:
            sku = ProductSku(
                tenant_id=1,
                workspace_id=1,
                product_id=product.id,
                name="黑色",
                code="black",
                keywords=["黑色"],
                image_asset_id=image.id,
                is_enabled=True,
            )
        batch = StandardDetailBatch(
            tenant_id=1,
            workspace_id=1,
            source_type="order_rows",
            status="ready",
        )
        db.add_all([sku, batch])
        db.flush()
        detail = StandardDetail(
            tenant_id=1,
            workspace_id=1,
            standard_detail_batch_id=batch.id,
            field_values={
                "capture_task_id": 8202,
                "product": "鞋",
                "sales_attr1": "黑色",
                "sales_attr2": "42",
                "quantity": "2",
                "remark": "加急",
            },
            image_match_status="pending",
            stall_match_status="pending",
        )
        db.add(detail)
        db.commit()
        return product.id, sku.id, image.id, detail.id


def _activate_test_recognition_pack() -> None:
    with SessionLocal() as db:
        db.query(RecognitionRulePack).filter(RecognitionRulePack.workspace_id == 1).update(
            {"is_enabled": False, "status": "inactive"}
        )
        payload = {
            "contract_version": "recognition_rule_pack_v1",
            "pack": {"code": "test-shoe-pack", "name": "测试鞋类面单规则", "version": "1.0.0"},
            "parser_policy": {"requires_active_rule_pack": True, "order_row_parser": "shoe_waybill_v1"},
        }
        pack = db.query(RecognitionRulePack).filter(
            RecognitionRulePack.workspace_id == 1,
            RecognitionRulePack.code == "test-shoe-pack",
            RecognitionRulePack.is_deleted.is_(False),
        ).first()
        if pack is None:
            pack = RecognitionRulePack(
                tenant_id=1,
                workspace_id=1,
                name="测试鞋类面单规则",
                code="test-shoe-pack",
                version="1.0.0",
                payload=payload,
                status="active",
                is_enabled=True,
            )
            db.add(pack)
        else:
            pack.name = "测试鞋类面单规则"
            pack.version = "1.0.0"
            pack.payload = payload
            pack.status = "active"
            pack.is_enabled = True
        db.commit()


def _deactivate_test_recognition_packs() -> None:
    with SessionLocal() as db:
        db.query(RecognitionRulePack).filter(RecognitionRulePack.workspace_id == 1).update(
            {"is_enabled": False, "status": "inactive"}
        )
        db.commit()


def test_product_matching_scoped_preview_requires_active_rule_pack() -> None:
    with TestClient(app) as client:
        headers = _headers(client)
        _deactivate_test_recognition_packs()
        _product_id, _sku_id, _image_id, detail_id = _seed_assets_and_detail()

        for scope in (
            {"scope_type": "selected_records", "standard_detail_ids": [detail_id], "confirmed_by_user": True},
            {"scope_type": "global", "confirmed_by_user": True},
        ):
            preview_response = client.post(
                "/api/v1/product-sku-linking/preview",
                headers=headers,
                json={"scope": scope, "include_saved_rules": True},
            )

            assert preview_response.status_code == 200
            body = preview_response.json()
            assert body["status"] == "rule_pack_missing"
            assert body["rule_pack_required"] is True
            assert body["summary"]["total"] == 0
            assert body["rows"] == []

        apply_response = client.post(
            "/api/v1/product-sku-linking/apply",
            headers=headers,
            json={
                "scope": {"scope_type": "selected_records", "standard_detail_ids": [detail_id], "confirmed_by_user": True},
                "include_enabled_rules": True,
            },
        )

        assert apply_response.status_code == 200
        apply_body = apply_response.json()
        assert apply_body["status"] == "rule_pack_missing"
        assert apply_body["rule_pack_required"] is True
        assert apply_body["applied_item_count"] == 0
        assert apply_body["summary"]["total"] == 0


def test_product_matching_preview_requires_explicit_batch_or_rows() -> None:
    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        with SessionLocal() as db:
            batch = StandardDetailBatch(
                tenant_id=1,
                workspace_id=1,
                source_type="legacy-standard-detail",
                status="ready",
            )
            db.add(batch)
            db.flush()
            db.add(
                StandardDetail(
                    tenant_id=1,
                    workspace_id=1,
                    standard_detail_batch_id=batch.id,
                    field_values={
                        "product": "这条旧详情不允许被全局商品匹配扫描",
                        "sales_attr1": "黑色",
                        "sales_attr2": "42",
                        "quantity": "1",
                        "remark": "",
                    },
                    image_match_status="pending",
                    stall_match_status="pending",
                )
            )
            db.commit()

        preview_response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={"scope": {"scope_type": "global", "confirmed_by_user": True}, "include_saved_rules": True},
        )

        assert preview_response.status_code == 200
        body = preview_response.json()
        assert body["status"] == "scope_required"
        assert body["summary"]["total"] == 0
        assert body["rows"] == []
        assert "采集批次" in body["message"]


def test_current_batch_preview_consumes_clean_order_rows_not_legacy_standard_detail_text(monkeypatch) -> None:
    _use_parser_service_stub(monkeypatch)

    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        with SessionLocal() as db:
            task = CaptureTask(
                tenant_id=1,
                workspace_id=1,
                name="测试采集任务",
                status="completed",
            )
            product = Product(
                tenant_id=1,
                workspace_id=1,
                name="反光滑板鞋",
                code="reflective-shoe",
                keywords=["新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C"],
                is_enabled=True,
            )
            batch = StandardDetailBatch(
                tenant_id=1,
                workspace_id=1,
                source_type="collector",
                status="ready",
            )
            db.add_all([task, product, batch])
            db.flush()
            dirty_text = "【新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C 158369C】黑色反光一星158477C 44(偏大一码，建议拍小) 1 件"
            detail = StandardDetail(
                tenant_id=1,
                workspace_id=1,
                standard_detail_batch_id=batch.id,
                field_values={
                    "capture_task_id": task.id,
                    "raw_record_id": 34,
                    "source_component": "cainiao-cnprint",
                    "source_index": "34",
                    "product_short_text": dirty_text,
                    "product_full_text": dirty_text,
                    "product_count_text": "1件",
                },
                image_match_status="pending",
                stall_match_status="pending",
            )
            db.add(detail)
            db.commit()
            task_id = task.id
            product_id = product.id

        response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={
                "scope": {"scope_type": "current_batch", "task_id": task_id, "confirmed_by_user": True},
                "rule": {
                    "product_id": product_id,
                    "product_match_fields": ["product"],
                    "product_keyword": "新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C 158369C",
                    "product_match_type": "contains",
                    "sku_match_fields": [],
                },
            },
        )

        assert response.status_code == 200
        body = response.json()
        assert body["summary"]["total"] == 1
        row_input = body["rows"][0]["input"]
        assert row_input["product"] == "新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C 158369C"
        assert row_input["sales_attr1"] == "黑色反光一星158477C"
        assert row_input["sales_attr2"] == "44"
        assert row_input["quantity"] == "1"


def test_report_preview_consumes_clean_order_rows_not_stale_product_matching_payloads(monkeypatch) -> None:
    _use_parser_service_stub(monkeypatch)

    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        with SessionLocal() as db:
            suffix = uuid.uuid4().hex[:8]
            task = CaptureTask(
                tenant_id=1,
                workspace_id=1,
                name=f"测试报货预览任务-{suffix}",
                status="completed",
            )
            product = Product(
                tenant_id=1,
                workspace_id=1,
                name=f"反光滑板鞋-{suffix}",
                code=f"reflective-shoe-report-{suffix}",
                keywords=["新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C"],
                is_enabled=True,
            )
            batch = StandardDetailBatch(
                tenant_id=1,
                workspace_id=1,
                source_type="collector",
                status="ready",
            )
            db.add_all([task, product, batch])
            db.flush()
            dirty_text = "【新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C 158369C】黑色反光一星158477C 44(偏大一码，建议拍小) 1 件"
            detail = StandardDetail(
                tenant_id=1,
                workspace_id=1,
                standard_detail_batch_id=batch.id,
                field_values={
                    "capture_task_id": task.id,
                    "raw_record_id": 34,
                    "source_component": "cainiao-cnprint",
                    "source_index": "34",
                    "product_short_text": dirty_text,
                    "product_full_text": dirty_text,
                    "product_count_text": "1件",
                    "product_sku_linking_results": [
                        {
                            "contract": "product-sku-linking-results-v1",
                            "standard_fields": {
                                "product": "建议拍小) 1 件",
                                "sales_attr1": "1件",
                                "sales_attr2": dirty_text,
                                "quantity": dirty_text.split("，")[0],
                                "remark": "",
                            },
                            "match_status": "product_unmatched",
                            "exception_reason": "旧写回结果，不应该污染新报货预览。",
                        }
                    ],
                },
                image_match_status="pending",
                stall_match_status="pending",
            )
            db.add(detail)
            db.add(
                ProductMatchingRule(
                    tenant_id=1,
                    workspace_id=1,
                    name="反光滑板鞋规则",
                    scope_type="global",
                    scope_payload={"scope_type": "global"},
                    product_id=product.id,
                    product_match_fields=["product"],
                    product_keyword="新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C 158369C",
                    product_match_type="contains",
                    sku_match_fields=[],
                    source_samples=[],
                    field_sources={},
                    preview_summary={},
                    is_enabled=True,
                )
            )
            db.commit()
            task_id = task.id

        response = client.get(f"/api/v1/collector-control/tasks/{task_id}/report-preview", headers=headers)

        assert response.status_code == 200
        body = response.json()
        assert body["summary"]["total"] == 1
        assert "fallback_matched" not in body["summary"]
        row = body["rows"][0]
        assert row["product_text"] == "新款特价黑色反光一星木村Low麂皮男鞋女鞋滑板鞋158477C 158369C"
        assert row["sales_attr1_text"] == "黑色反光一星158477C"
        assert row["sales_attr2_text"] == "44"
        assert row["quantity_text"] == "1"


def test_product_matching_rule_preview_save_apply_and_disable() -> None:
    with TestClient(app) as client:
        headers = _headers(client)
        product_id, sku_id, image_id, detail_id = _seed_assets_and_detail()
        scope = {
            "scope_type": "selected_records",
            "standard_detail_ids": [detail_id],
            "confirmed_by_user": True,
        }
        draft_rule = {
            "name": "鞋类黑色",
            "scope": scope,
            "product_id": product_id,
            "product_match_fields": ["product"],
            "product_keyword": "鞋",
            "product_match_type": "contains",
            "sku_match_fields": ["sales_attr1"],
            "sku_id": sku_id,
            "image_asset_id": image_id,
            "source_samples": [
                {
                    "product": "鞋",
                    "sales_attr1": "黑色",
                    "sales_attr2": "42",
                    "quantity": "2",
                    "remark": "加急",
                    "raw_payload": "不能保存到商品匹配学习记录来源样本",
                    "productInfo": "不能保存到商品匹配学习记录来源样本",
                }
            ],
            "field_sources": {"product": "面单解析后的五字段.product"},
            "revision_note": "用户确认鞋类黑色规则",
        }

        with SessionLocal() as db:
            rule_count_before_preview = db.query(ProductMatchingRule).count()

        preview_response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={"scope": scope, "rule": draft_rule},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["summary"]["matched"] == 1
        assert preview["summary"]["product_unmatched"] == 0
        assert preview["summary"]["sku_unmatched"] == 0
        assert preview["summary"]["image_unmatched"] == 0
        assert preview["summary"]["conflict"] == 0
        assert preview["samples"]["matched"][0]["input"]["product"] == "鞋"

        empty_selected_preview = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={
                "scope": {"scope_type": "selected_records", "standard_detail_ids": [], "confirmed_by_user": True},
                "rule": draft_rule,
            },
        )
        assert empty_selected_preview.status_code == 200
        assert empty_selected_preview.json()["summary"]["total"] == 0

        with SessionLocal() as db:
            detail = db.get(StandardDetail, detail_id)
            assert detail is not None
            assert "product_sku_linking_results" not in detail.field_values
            assert db.query(ProductMatchingRule).count() == rule_count_before_preview

        save_response = client.post("/api/v1/product-sku-linking/rules", headers=headers, json=draft_rule)
        assert save_response.status_code == 201
        saved_rule = save_response.json()["rule"]
        assert saved_rule["product_keyword"] == "鞋"
        assert saved_rule["scope_type"] == "global"
        assert saved_rule["scope_payload"]["scope_type"] == "global"
        assert saved_rule["preview_summary"]["matched"] == 1
        assert saved_rule["source_samples"] == [
            {
                "product": "鞋",
                "sales_attr1": "黑色",
                "sales_attr2": "42",
                "quantity": "2",
                "remark": "加急",
            }
        ]

        rules_response = client.get("/api/v1/product-sku-linking/rules", headers=headers)
        assert rules_response.status_code == 200
        assert any(rule["id"] == saved_rule["id"] for rule in rules_response.json()["rules"])

        revise_response = client.patch(
            f"/api/v1/product-sku-linking/rules/{saved_rule['id']}",
            headers=headers,
            json={"product_keyword": "鞋", "revision_note": "修订测试"},
        )
        assert revise_response.status_code == 200
        assert revise_response.json()["rule"]["revision"] == saved_rule["revision"] + 1

        empty_selected_apply = client.post(
            "/api/v1/product-sku-linking/apply",
            headers=headers,
            json={
                "scope": {"scope_type": "selected_records", "standard_detail_ids": [], "confirmed_by_user": True},
                "include_enabled_rules": True,
            },
        )
        assert empty_selected_apply.status_code == 200
        empty_apply_body = empty_selected_apply.json()
        assert empty_apply_body["summary"]["total"] == 0
        assert empty_apply_body["applied_detail_count"] == 0
        assert empty_apply_body["applied_item_count"] == 0

        with SessionLocal() as db:
            detail = db.get(StandardDetail, detail_id)
            assert detail is not None
            assert "product_sku_linking_results" not in detail.field_values

        apply_response = client.post(
            "/api/v1/product-sku-linking/apply",
            headers=headers,
            json={"scope": scope, "rule_ids": [saved_rule["id"]], "include_enabled_rules": True},
        )
        assert apply_response.status_code == 200
        assert apply_response.json()["applied_detail_count"] == 1
        assert apply_response.json()["summary"]["matched"] == 1

        with SessionLocal() as db:
            detail = db.get(StandardDetail, detail_id)
            assert detail is not None
            results = detail.field_values["product_sku_linking_results"]
            assert results[0]["contract"] == "product-sku-linking-results-v1"
            assert results[0]["product"] == "帆布鞋"
            assert results[0]["sku"] == "黑色"
            assert results[0]["image_asset_id"] == image_id
            assert results[0]["match_status"] == "matched"

        disable_response = client.patch(
            f"/api/v1/product-sku-linking/rules/{saved_rule['id']}",
            headers=headers,
            json={"is_enabled": False, "revision_note": "停用测试"},
        )
        assert disable_response.status_code == 200
        assert disable_response.json()["rule"]["is_enabled"] is False

        disabled_preview = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={"scope": scope, "include_saved_rules": True},
        )
        assert disabled_preview.status_code == 200
        disabled_preview_body = disabled_preview.json()
        assert disabled_preview_body["summary"]["matched"] == 0
        assert disabled_preview_body["summary"]["product_unmatched"] == 1
        assert disabled_preview_body["samples"]["product_unmatched"][0]["match_source"] == "user_learning_rule"

        delete_response = client.delete(f"/api/v1/product-sku-linking/rules/{saved_rule['id']}", headers=headers)
        assert delete_response.status_code == 204

        deleted_rules_response = client.get("/api/v1/product-sku-linking/rules?include_disabled=true", headers=headers)
        assert deleted_rules_response.status_code == 200
        assert all(rule["id"] != saved_rule["id"] for rule in deleted_rules_response.json()["rules"])


def _seed_order_row_draft_source_for_product_matching(suffix: str = "") -> tuple[int, int, int, int, int]:
    suffix = suffix or f"-{uuid.uuid4().hex[:8]}"
    with SessionLocal() as db:
        task = CaptureTask(
            tenant_id=1,
            workspace_id=1,
            name=f"订单行商品匹配测试{suffix}",
            status="completed",
        )
        product = Product(
            tenant_id=1,
            workspace_id=1,
            name=f"户外登山鞋{suffix}",
            code=f"outdoor-shoe{suffix}",
            keywords=["登山鞋"],
            is_enabled=True,
        )
        image = ImageAsset(
            tenant_id=1,
            workspace_id=1,
            name=f"户外鞋图{suffix}",
            file_path="storage/outdoor.png",
        )
        db.add_all([task, product, image])
        db.flush()
        sku = ProductSku(
            tenant_id=1,
            workspace_id=1,
            product_id=product.id,
            name="低帮",
            code="low",
            keywords=["低帮"],
            image_asset_id=image.id,
            is_enabled=True,
        )
        batch = StandardDetailBatch(
            tenant_id=1,
            workspace_id=1,
            source_type="collector",
            status="ready",
        )
        product_line = (
            "【登山鞋涉水鞋防水机能户外徒步男鞋女鞋休闲防滑溯溪越野鞋跑步鞋】低帮黑白 42 1件；"
            "【登山鞋涉水鞋防水机能户外徒步男鞋女鞋休闲防滑溯溪越野鞋跑步鞋】低帮浅绿 36 1件"
        )
        raw_record = RawCaptureRecord(
            tenant_id=1,
            workspace_id=1,
            task_id=task.id,
            source_component="cloud-print-client",
            source_index=f"multi{suffix}",
            raw_payload=json.dumps(
                {
                    "task": {
                        "documents": [
                            {
                                "documentID": f"MULTI{suffix}",
                                "contents": [
                                    {
                                        "data": {
                                            "productCount": "2件",
                                            "productInfo": product_line,
                                            "productShortInfo": product_line,
                                        }
                                    }
                                ],
                            }
                        ]
                    }
                },
                ensure_ascii=False,
            ),
            payload_format="json",
            status="parsed",
        )
        db.add_all([sku, batch, raw_record])
        db.flush()
        detail = StandardDetail(
            tenant_id=1,
            workspace_id=1,
            standard_detail_batch_id=batch.id,
            field_values={
                "capture_task_id": task.id,
                "raw_record_id": raw_record.id,
                "source_component": "cloud-print-client",
                "source_index": "24",
                "product_short_text": product_line,
                "product_full_text": product_line,
                "product_count_text": "2件",
            },
            image_match_status="pending",
            stall_match_status="pending",
        )
        db.add(detail)
        db.commit()
        return task.id, product.id, sku.id, image.id, detail.id


def _add_unmapped_raw_record(task_id: int) -> int:
    with SessionLocal() as db:
        raw_record = RawCaptureRecord(
            tenant_id=1,
            workspace_id=1,
            task_id=task_id,
            source_component="cloud-print-client",
            raw_payload='{"productInfo":"还没有字段定义的面单"}',
            payload_format="json",
            status="parsed",
        )
        db.add(raw_record)
        db.commit()
        return raw_record.id


def _add_unmapped_multi_sample_raw_record(task_id: int) -> int:
    payload = {
        "task": {
            "documents": [
                {
                    "documentID": "UNMAPPED-1",
                    "contents": [{"printXML": "<text><![CDATA[未定义商品A*1]]></text>"}],
                },
                {
                    "documentID": "UNMAPPED-2",
                    "contents": [{"printXML": "<text><![CDATA[未定义商品B*1]]></text>"}],
                },
            ]
        }
    }
    with SessionLocal() as db:
        raw_record = RawCaptureRecord(
            tenant_id=1,
            workspace_id=1,
            task_id=task_id,
            source_component="cloud-print-client",
            raw_payload=json.dumps(payload, ensure_ascii=False),
            payload_format="json",
            status="parsed",
        )
        db.add(raw_record)
        db.commit()
        return raw_record.id


def test_product_matching_consumes_order_row_drafts_and_export_uses_clean_results(monkeypatch) -> None:
    _use_parser_service_stub(monkeypatch)

    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        task_id, product_id, sku_id, image_id, detail_id = _seed_order_row_draft_source_for_product_matching()
        scope = {
            "scope_type": "current_batch",
            "task_id": task_id,
            "confirmed_by_user": True,
        }
        draft_rule = {
            "name": "登山鞋多商品",
            "scope": scope,
            "product_id": product_id,
            "product_match_fields": ["product"],
            "product_keyword": "登山鞋",
            "product_match_type": "contains",
            "sku_match_fields": ["sales_attr1"],
            "sku_id": sku_id,
            "image_asset_id": image_id,
        }

        preview_response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={"scope": scope, "rule": draft_rule},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["summary"]["total"] == 2
        assert preview["summary"]["matched"] == 2

        save_response = client.post("/api/v1/product-sku-linking/rules", headers=headers, json=draft_rule)
        assert save_response.status_code == 201
        saved_rule_id = save_response.json()["rule"]["id"]

        saved_preview_response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={"scope": scope, "include_saved_rules": True, "rule_ids": [saved_rule_id]},
        )
        assert saved_preview_response.status_code == 200
        saved_preview = saved_preview_response.json()
        assert saved_preview["summary"]["matched"] == 2
        assert saved_preview["samples"]["matched"][0]["matched_linking_rule"]["id"] == saved_rule_id

        apply_response = client.post(
            "/api/v1/product-sku-linking/apply",
            headers=headers,
            json={"scope": scope, "rule_ids": [saved_rule_id], "include_enabled_rules": True},
        )
        assert apply_response.status_code == 200
        body = apply_response.json()
        assert body["applied_standard_detail_count"] == 1
        assert body["applied_item_count"] == 2

        with SessionLocal() as db:
            detail = db.get(StandardDetail, detail_id)
            assert detail is not None
            payloads = detail.field_values["product_sku_linking_results"]
            assert len(payloads) == 2
            assert payloads[0]["standard_fields"]["sales_attr1"] == "低帮黑白"
            assert payloads[1]["standard_fields"]["sales_attr2"] == "36"

        export_preview = client.get(f"/api/v1/collector-control/tasks/{task_id}/recognition-preview", headers=headers)
        assert export_preview.status_code == 200
        export_body = export_preview.json()
        assert export_body["data_source"] == "order_row_drafts"
        assert len(export_body["rows"]) == 2
        assert export_body["rows"][0]["sales_attr1_text"] == "低帮黑白"
        assert export_body["rows"][1]["sales_attr2_text"] == "36"

        workbook_response = client.get(
            f"/api/v1/collector-control/tasks/{task_id}/report-workbook",
            headers=headers,
            params={"layout": json.dumps({"output_mode": "merged_sheet"}, ensure_ascii=False)},
        )
        assert workbook_response.status_code == 200
        disposition = unquote(workbook_response.headers["content-disposition"])
        assert "capture-task" not in disposition
        assert f"-{task_id}-" not in disposition
        assert re.search(r"订单整理文档_\d{8}_\d{6}\.xlsx", disposition)
        workbook = load_workbook(BytesIO(workbook_response.content))
        assert workbook.sheetnames == ["报货表", "异常面单"]

        report_sheet = workbook["报货表"]
        assert [cell.value for cell in report_sheet[1]] == [
            "商品",
            "销售属性1",
            "图片",
            "销售属性2",
            "数量",
            "备注",
            "图片匹配文本",
        ]
        assert report_sheet.max_row == 3
        assert str(report_sheet.cell(row=2, column=1).value).startswith("户外登山鞋")
        exported_items = {
            (
                report_sheet.cell(row=row, column=2).value,
                report_sheet.cell(row=row, column=4).value,
                report_sheet.cell(row=row, column=5).value,
            )
            for row in range(2, 4)
        }
        assert exported_items == {("低帮黑白", "42", 1), ("低帮浅绿", "36", 1)}
        assert all(report_sheet.cell(row=row, column=3).value is None for row in range(2, 4))

        exception_sheet = workbook["异常面单"]
        assert [cell.value for cell in exception_sheet[1]] == ["图片匹配文本"]
        assert exception_sheet.max_row == 1


def test_product_matching_preview_reports_batch_order_row_coverage(monkeypatch) -> None:
    _use_parser_service_stub(monkeypatch)

    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        task_id, product_id, _sku_id, _image_id, _detail_id = _seed_order_row_draft_source_for_product_matching(
            "-coverage"
        )
        _add_unmapped_multi_sample_raw_record(task_id)
        scope = {
            "scope_type": "current_batch",
            "task_id": task_id,
            "confirmed_by_user": True,
        }
        draft_rule = {
            "name": "登山鞋覆盖率",
            "scope": scope,
            "product_id": product_id,
            "product_match_fields": ["product"],
            "product_keyword": "登山鞋",
            "product_match_type": "contains",
        }

        preview_response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={"scope": scope, "rule": draft_rule},
        )

        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["summary"]["total"] == 4
        assert preview["coverage"]["task_id"] == task_id
        assert preview["coverage"]["total_raw_record_count"] == 2
        assert preview["coverage"]["total_waybill_count"] == 3
        assert preview["coverage"]["order_row_waybill_count"] == 3
        assert preview["coverage"]["standard_row_count"] == 4
        assert preview["coverage"]["missing_order_row_count"] == 0
        assert preview["coverage"]["source_type_counts"] == {"order_row_draft": 4}


def test_recognition_preview_uses_waybill_sequence_not_raw_record_source_index(monkeypatch) -> None:
    _use_parser_service_stub(monkeypatch)

    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        with SessionLocal() as db:
            task = CaptureTask(
                tenant_id=1,
                workspace_id=1,
                name="异常页编号口径测试",
                status="completed",
            )
            db.add(task)
            db.flush()
            for source_index, product_text in (
                ("2648", "2025新款网面女鞋男鞋情侣透气跑步鞋 5.0二代灰色 38.5 1件"),
                ("7132", "范33 带木one帆布kw 木村-3M反光 42.5 1件"),
            ):
                payload = {
                    "task": {
                        "documents": [
                            {
                                "documentID": f"DOC-{source_index}",
                                "contents": [
                                    {
                                        "data": {
                                            "productCount": "1件",
                                            "productInfo": product_text,
                                            "productShortInfo": product_text,
                                        }
                                    }
                                ],
                            }
                        ]
                    }
                }
                db.add(
                    RawCaptureRecord(
                        tenant_id=1,
                        workspace_id=1,
                        task_id=task.id,
                        source_component="cainiao-cnprint",
                        source_index=source_index,
                        raw_payload=json.dumps(payload, ensure_ascii=False),
                        payload_format="json",
                        status="parsed",
                    )
                )
            db.commit()
            task_id = task.id

        response = client.get(f"/api/v1/collector-control/tasks/{task_id}/recognition-preview", headers=headers)

        assert response.status_code == 200
        body = response.json()
        assert body["detail_count"] == 2
        assert body["waybill_count"] == 2
        assert body["order_row_count"] == 2
        labels = [row["source_label"] for row in body["rows"]]
        assert labels == ["第1批-第1单-子1", "第1批-第2单-子1"]
        assert all("2648" not in label and "7132" not in label for label in labels)


def test_current_batch_product_matching_uses_parser_service_order_rows(monkeypatch) -> None:
    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        with SessionLocal() as db:
            task = CaptureTask(
                tenant_id=1,
                workspace_id=1,
                name="同源解析服务商品匹配测试",
                status="completed",
            )
            batch = StandardDetailBatch(
                tenant_id=1,
                workspace_id=1,
                source_type="collector",
                status="ready",
            )
            db.add_all([task, batch])
            db.flush()
            detail = StandardDetail(
                tenant_id=1,
                workspace_id=1,
                standard_detail_batch_id=batch.id,
                field_values={
                    "capture_task_id": task.id,
                    "raw_record_id": 999001,
                    "source_component": "cloud-print-client",
                    "source_index": "remote-service-sample",
                    "product_short_text": "本地旧解析不应该被商品匹配使用 99 1件",
                    "product_count_text": "1件",
                },
                image_match_status="pending",
                stall_match_status="pending",
            )
            db.add(detail)
            db.commit()
            task_id = task.id

        def fake_parse_order_row_drafts_with_service(**kwargs):
            assert kwargs["task_id"] == task_id
            return {
                "contract_version": "order_row_drafts_v1",
                "task_id": task_id,
                "summary": {
                    "parent_waybill_count": 1,
                    "child_waybill_count": 1,
                    "draft_count": 1,
                    "needs_review_count": 0,
                    "special_count": 0,
                },
                "parents": [
                    {
                        "raw_record_id": 999001,
                        "task_id": task_id,
                        "parent_label": "远端解析父单",
                        "source_component": "cloud-print-client",
                        "source_index": "remote-service-sample",
                        "child_count": 1,
                        "rows": [],
                    }
                ],
                "rows": [
                    {
                        "raw_record_id": 999001,
                        "task_id": task_id,
                        "parent_label": "远端解析父单",
                        "child_label": "远端解析子单",
                        "child_index": 1,
                        "child_count": 1,
                        "source_component": "cloud-print-client",
                        "source_index": "remote-service-sample",
                        "product": "远端解析干净商品",
                        "sales_attr1": "远端颜色",
                        "sales_attr2": "42",
                        "quantity": 1,
                        "remark": "",
                        "image_match_text": "远端解析干净商品 远端颜色 42 1",
                        "original_text": "远端解析原文",
                        "status": "draft",
                        "review_reason": "",
                    }
                ],
            }

        monkeypatch.setattr(order_row_reader_service, "waybill_parser_service_enabled", lambda: True, raising=False)
        monkeypatch.setattr(
            order_row_reader_service,
            "parse_order_row_drafts_with_service",
            fake_parse_order_row_drafts_with_service,
            raising=False,
        )

        response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={
                "scope": {"scope_type": "current_batch", "task_id": task_id, "confirmed_by_user": True},
                "include_saved_rules": True,
            },
        )

        assert response.status_code == 200
        body = response.json()
        assert body["summary"]["total"] == 1
        assert body["rows"][0]["input"] == {
            "product": "远端解析干净商品",
            "sales_attr1": "远端颜色",
            "sales_attr2": "42",
            "quantity": "1",
            "remark": "",
        }


def test_current_batch_product_matching_prefers_current_raw_samples_over_stale_standard_details(
    monkeypatch,
) -> None:
    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        payload = {
            "task": {
                "documents": [
                    {
                        "documentID": "RAW-MATCH-CURRENT",
                        "contents": [
                            {
                                "data": {
                                    "productCount": "1件",
                                    "productInfo": "当前解析商品 蓝色 40 1件",
                                    "productShortInfo": "当前解析商品 蓝色 40 1件",
                                }
                            }
                        ],
                    }
                ]
            }
        }
        with SessionLocal() as db:
            task = CaptureTask(
                tenant_id=1,
                workspace_id=1,
                name="商品匹配同源数据测试",
                status="completed",
            )
            batch = StandardDetailBatch(
                tenant_id=1,
                workspace_id=1,
                source_type="collector",
                status="ready",
            )
            db.add_all([task, batch])
            db.flush()
            db.add(
                RawCaptureRecord(
                    tenant_id=1,
                    workspace_id=1,
                    task_id=task.id,
                    document_id="RAW-MATCH-CURRENT",
                    source_component="cloud-print-client",
                    source_index="current-match-raw",
                    payload_format="json",
                    raw_payload=json.dumps(payload, ensure_ascii=False),
                    status="parsed",
                )
            )
            db.add(
                StandardDetail(
                    tenant_id=1,
                    workspace_id=1,
                    standard_detail_batch_id=batch.id,
                    field_values={
                        "capture_task_id": task.id,
                        "raw_record_id": 880088,
                        "source_component": "legacy-cache",
                        "source_index": "stale-match-cache",
                        "product_short_text": "旧缓存商品 黑色 39 1件",
                    },
                    image_match_status="pending",
                    stall_match_status="pending",
                )
            )
            db.commit()
            task_id = task.id

        def fake_parse_order_row_drafts_with_service(**kwargs):
            assert kwargs["task_id"] == task_id
            assert kwargs.get("standard_details") == []
            assert kwargs.get("waybill_samples")
            return {
                "contract_version": "order_row_drafts_v1",
                "task_id": task_id,
                "summary": {
                    "parent_waybill_count": 1,
                    "child_waybill_count": 1,
                    "draft_count": 1,
                    "needs_review_count": 0,
                    "special_count": 0,
                },
                "parents": [
                    {
                        "raw_record_id": 880077,
                        "task_id": task_id,
                        "parent_label": "第1批-第1单",
                        "source_component": "cloud-print-client",
                        "source_index": "current-match-raw",
                        "child_count": 1,
                        "rows": [],
                    }
                ],
                "rows": [
                    {
                        "raw_record_id": 880077,
                        "task_id": task_id,
                        "parent_label": "第1批-第1单",
                        "child_label": "第1批-第1单-子1",
                        "child_index": 1,
                        "child_count": 1,
                        "source_component": "cloud-print-client",
                        "source_index": "current-match-raw",
                        "product": "当前解析商品",
                        "sales_attr1": "蓝色",
                        "sales_attr2": "40",
                        "quantity": 1,
                        "remark": "",
                        "image_match_text": "当前解析商品 蓝色 40 1",
                        "original_text": "当前解析原文",
                        "status": "draft",
                        "review_reason": "",
                    }
                ],
            }

        monkeypatch.setattr(order_row_reader_service, "waybill_parser_service_enabled", lambda: True, raising=False)
        monkeypatch.setattr(
            order_row_reader_service,
            "parse_order_row_drafts_with_service",
            fake_parse_order_row_drafts_with_service,
            raising=False,
        )

        response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={
                "scope": {"scope_type": "current_batch", "task_id": task_id, "confirmed_by_user": True},
                "include_saved_rules": True,
            },
        )

        assert response.status_code == 200
        body = response.json()
        assert body["summary"]["total"] == 1
        assert body["rows"][0]["input"]["product"] == "当前解析商品"
        assert "旧缓存商品" not in json.dumps(body, ensure_ascii=False)


def test_current_batch_product_matching_reports_parser_service_failure(monkeypatch) -> None:
    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        with SessionLocal() as db:
            task = CaptureTask(
                tenant_id=1,
                workspace_id=1,
                name="解析服务失败测试",
                status="completed",
            )
            batch = StandardDetailBatch(
                tenant_id=1,
                workspace_id=1,
                source_type="collector",
                status="ready",
            )
            db.add_all([task, batch])
            db.flush()
            db.add(
                StandardDetail(
                    tenant_id=1,
                    workspace_id=1,
                    standard_detail_batch_id=batch.id,
                    field_values={
                        "capture_task_id": task.id,
                        "raw_record_id": 990101,
                        "source_component": "cloud-print-client",
                        "source_index": "parser-failure-sample",
                        "product_short_text": "本地旧解析不允许兜底使用 42 1件",
                        "product_count_text": "1件",
                    },
                    image_match_status="pending",
                    stall_match_status="pending",
                )
            )
            db.commit()
            task_id = task.id

        def fail_parse_order_row_drafts_with_service(**_kwargs):
            raise RuntimeError("parser service offline")

        monkeypatch.setattr(order_row_reader_service, "waybill_parser_service_enabled", lambda: True, raising=False)
        monkeypatch.setattr(
            order_row_reader_service,
            "parse_order_row_drafts_with_service",
            fail_parse_order_row_drafts_with_service,
            raising=False,
        )

        response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={
                "scope": {"scope_type": "current_batch", "task_id": task_id, "confirmed_by_user": True},
                "include_saved_rules": True,
            },
        )

        assert response.status_code == 502
        assert "面单解析服务" in response.json()["detail"]


def test_current_batch_product_matching_requires_parser_service(monkeypatch) -> None:
    with TestClient(app) as client:
        headers = _headers(client)
        _activate_test_recognition_pack()
        with SessionLocal() as db:
            task = CaptureTask(
                tenant_id=1,
                workspace_id=1,
                name="解析服务必须启用测试",
                status="completed",
            )
            batch = StandardDetailBatch(
                tenant_id=1,
                workspace_id=1,
                source_type="collector",
                status="ready",
            )
            db.add_all([task, batch])
            db.flush()
            db.add(
                StandardDetail(
                    tenant_id=1,
                    workspace_id=1,
                    standard_detail_batch_id=batch.id,
                    field_values={
                        "capture_task_id": task.id,
                        "raw_record_id": 990202,
                        "source_component": "cloud-print-client",
                        "source_index": "parser-service-required",
                        "product_short_text": "旧解析不能继续兜底 42 1件",
                        "product_count_text": "1件",
                    },
                    image_match_status="pending",
                    stall_match_status="pending",
                )
            )
            db.commit()
            task_id = task.id

        monkeypatch.setattr(order_row_reader_service, "waybill_parser_service_enabled", lambda: False, raising=False)

        response = client.post(
            "/api/v1/product-sku-linking/preview",
            headers=headers,
            json={
                "scope": {"scope_type": "current_batch", "task_id": task_id, "confirmed_by_user": True},
                "include_saved_rules": True,
            },
        )

        assert response.status_code == 503
        assert "旧解析兜底" in response.json()["detail"]
