from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from app.api.routes.collector_runtime import (
    RECOGNITION_EXCEPTION_SHEET_TITLE,
    append_recognition_exception_sheet,
    append_xlsx_rows,
    pending_unmapped_waybill_product_sku_linking_row,
    product_sku_linking_export_row,
    recognition_exception_export_rows,
    recognition_report_export_rows,
    recognition_report_headers,
    recognition_report_line_items,
    recognition_report_row_is_exportable,
    recognition_report_workbook,
    recognition_report_rows_by_stall,
    recognition_rows_from_product_sku_linking_results,
    report_quantity_value,
)


def test_pending_unmapped_export_text_does_not_expose_batch_numbers() -> None:
    row = pending_unmapped_waybill_product_sku_linking_row(
        {"sample_id": "sample-1", "raw_record_id": 1, "sample_text": "鞋款文字 42 *1"},
        detail_number=7,
    )

    assert row["source_label"] == "第1批-第7单"
    assert row["image_match_text"] == "鞋款文字 42 *1"
    assert recognition_exception_export_rows([row]) == [["鞋款文字 42 *1"]]


def test_report_preview_keeps_matched_rule_id_for_exception_repair() -> None:
    row = product_sku_linking_export_row(
        {
            "match_status": "sku_ambiguous",
            "matched_rule": {"id": 35},
            "standard_fields": {"product": "秒45 按跑", "sales_attr1": "Cloudmonster Void灰色"},
        },
        source_identifiers={"detail_id": None},
        candidate_key_fallback="order-row:1",
        detail_number=1,
        item_index=1,
        item_count=1,
    )

    assert row["rule_id"] == 35


def test_export_contract_keeps_matched_rows_with_or_without_images() -> None:
    details = [
        SimpleNamespace(
            id=101,
            field_values={
                "product_sku_linking_result": {
                    "match_status": "matched",
                    "product": "鞋款A",
                    "product_id": 1,
                    "sku": "黑色42",
                    "image": {"id": 1, "name": "黑色图"},
                    "stall": {"id": 9, "name": "至尚"},
                    "standard_fields": {
                        "sales_attr1": "黑色",
                        "sales_attr2": "42",
                        "quantity": "2",
                        "remark": "加急",
                    },
                    "image_match_text": "鞋款A 黑色 42",
                },
            },
        ),
        SimpleNamespace(
            id=102,
            field_values={
                "product_sku_linking_result": {
                    "match_status": "matched",
                    "product": "鞋款B",
                    "product_id": 2,
                    "sku": "白色41",
                    "image": None,
                    "stall": {"id": 9, "name": "至尚"},
                    "standard_fields": {
                        "sales_attr1": "白色",
                        "sales_attr2": "41",
                        "quantity": "1",
                        "remark": "",
                    },
                    "image_match_text": "鞋款B 白色 41",
                },
            },
        ),
    ]

    rows = recognition_rows_from_product_sku_linking_results(details)

    assert rows[0]["status"] == "matched"
    assert rows[0]["stall_name"] == "至尚"
    assert rows[0]["stall_id"] == 9
    line_items = recognition_report_line_items(rows)
    assert line_items[0]["stall_name"] == "至尚"
    assert line_items[0]["image_label"] == "黑色图"
    assert line_items[1]["image_label"] == ""
    assert recognition_report_rows_by_stall(line_items)["至尚"][0]["product_category"] == "鞋款A"
    assert recognition_report_export_rows(rows) == [
        ["鞋款A", "黑色", "", "42", 2, "加急", "鞋款A 黑色 42"],
        ["鞋款B", "白色", "", "41", 1, "", "鞋款B 白色 41"],
    ]
    assert recognition_exception_export_rows(rows) == []


def test_export_routes_non_matched_and_special_rows_to_exception_sheet() -> None:
    details = [
        SimpleNamespace(
            id=102,
            field_values={
                "product_sku_linking_result": {
                    "match_status": "product_unmatched",
                    "standard_fields": {
                        "product": "未维护鞋款",
                        "sales_attr1": "蓝色",
                        "sales_attr2": "41",
                        "quantity": "1",
                        "remark": "",
                    },
                    "image_match_text": "未维护鞋款 蓝色 41",
                    "exception_reason": "商品未命中",
                },
            },
        ),
        SimpleNamespace(
            id=103,
            field_values={
                "product_sku_linking_result": {
                    "match_status": "special",
                    "standard_fields": {},
                    "image_match_text": "特殊面单原文",
                    "exception_reason": "特殊面单",
                },
            },
        ),
    ]

    rows = recognition_rows_from_product_sku_linking_results(details)

    assert recognition_report_export_rows(rows) == []
    assert recognition_exception_export_rows(rows) == [
        ["未维护鞋款 蓝色 41"],
        ["特殊面单原文"],
    ]


def test_normal_export_requires_a_real_matched_product() -> None:
    valid = {
        "status": "matched",
        "product_id": 10,
        "product_name": "鞋款A",
        "image_match_text": "鞋款A 黑色 42",
    }
    missing_id = {**valid, "product_id": None}
    missing_name = {**valid, "product_name": ""}

    assert recognition_report_row_is_exportable(valid) is True
    assert recognition_report_row_is_exportable(missing_id) is False
    assert recognition_report_row_is_exportable(missing_name) is False
    assert recognition_exception_export_rows([missing_id, missing_name]) == [
        ["鞋款A 黑色 42"],
        ["鞋款A 黑色 42"],
    ]


def test_normal_report_always_keeps_all_seven_business_columns() -> None:
    hidden_columns = [
        {"key": "product_name", "label": "商品", "visible": True},
        {"key": "sales_attr1", "label": "销售属性1", "visible": True},
        {"key": "sku_image", "label": "图片", "visible": True},
        {"key": "sales_attr2", "label": "销售属性2", "visible": True},
        {"key": "quantity", "label": "数量", "visible": True},
        {"key": "remark", "label": "备注", "visible": True},
        {"key": "image_match_text", "label": "图片匹配文本", "visible": False},
    ]

    assert recognition_report_headers({"columns": hidden_columns}) == [
        "商品",
        "销售属性1",
        "图片",
        "销售属性2",
        "数量",
        "备注",
        "图片匹配文本",
    ]


def test_normal_report_ignores_legacy_stall_column() -> None:
    layout = {
        "columns": [
            {"key": "stall_name", "label": "档口", "visible": True},
            {"key": "product_name", "label": "商品", "visible": True},
        ]
    }
    row = {
        "status": "matched",
        "product_id": 1,
        "product_name": "鞋款A",
        "stall_name": "至尚",
        "sales_attr1_text": "黑色",
        "sales_attr2_text": "42",
        "quantity_text": "1",
        "remark_text": "",
        "image_match_text": "鞋款A 黑色 42",
    }

    assert recognition_report_headers(layout) == [
        "商品",
        "销售属性1",
        "图片",
        "销售属性2",
        "数量",
        "备注",
        "图片匹配文本",
    ]
    assert recognition_report_export_rows([row], layout) == [
        ["鞋款A", "黑色", "", "42", 1, "", "鞋款A 黑色 42"],
    ]


def test_export_keeps_matched_rows_with_missing_sales_attrs_in_normal_sheet() -> None:
    details = [
        SimpleNamespace(
            id=103,
            field_values={
                "product_sku_linking_result": {
                    "match_status": "matched",
                    "product": "鞋款A",
                    "product_id": 1,
                    "sku": "黑色图",
                    "image": {"id": 1, "name": "黑色图"},
                    "standard_fields": {
                        "sales_attr1": "5.0二代灰色",
                        "sales_attr2": "",
                        "quantity": "1",
                        "remark": "",
                    },
                    "image_match_text": "鞋款A 5.0二代灰色",
                },
            },
        ),
    ]

    rows = recognition_rows_from_product_sku_linking_results(details)

    assert recognition_report_export_rows(rows) == [
        ["鞋款A", "5.0二代灰色", "", "-", 1, "", "鞋款A 5.0二代灰色"],
    ]
    assert recognition_exception_export_rows(rows) == []


def test_recognition_report_export_respects_sales_attr1_stack_layout() -> None:
    rows = [
        {
            "status": "matched",
            "product_name": "4.0",
            "product_id": 10,
            "sku_id": 20,
            "sku_image_asset_id": 30,
            "image_label": "4.0图",
            "sales_attr1_text": "二代灰白",
            "sales_attr2_text": "41",
            "quantity_text": "1",
            "item_count": 1,
            "remark_text": "",
            "image_match_text": "4.0 二代灰白 41",
        },
        {
            "status": "matched",
            "product_name": "4.0",
            "product_id": 10,
            "sku_id": 20,
            "sku_image_asset_id": 30,
            "image_label": "4.0图",
            "sales_attr1_text": "二代黑白",
            "sales_attr2_text": "42",
            "quantity_text": "1",
            "item_count": 1,
            "remark_text": "",
            "image_match_text": "4.0 二代黑白 42",
        },
    ]

    exported = recognition_report_export_rows(rows, {"stack_sales_attr1": True})

    assert len(exported) == 1
    assert exported[0][0] == "4.0"
    assert exported[0][1] == "二代灰白 二代黑白"
    assert exported[0][2] == ""
    assert exported[0][3] == "41 42"
    assert exported[0][4] == 2


def test_recognition_report_workbook_matches_preview_pixel_layout() -> None:
    report_rows = [
        {
            "product_category": "4.0",
            "stall_name": "1199",
            "spec": "二代灰白",
            "image_label": "4.0图",
            "sku_image_asset_id": 30,
            "size_text": "41",
            "quantity": 1,
            "remark_text": "",
            "image_match_text": "4.0 二代灰白 41",
        }
    ]
    layout = {
        "columns": [
            {"key": "product_name", "label": "商品", "visible": True, "width": 16},
            {"key": "sales_attr1", "label": "销售属性1", "visible": True, "width": 24},
            {"key": "sku_image", "label": "图片", "visible": True, "width": 18},
            {"key": "sales_attr2", "label": "销售属性2", "visible": True, "width": 18},
            {"key": "quantity", "label": "数量", "visible": True, "width": 12},
        ],
        "header_row_height": 32,
        "row_height": 120,
        "image_width": 96,
        "image_height": 96,
    }

    workbook = recognition_report_workbook(
        report_rows=report_rows,
        report_layout=layout,
        images_by_id={},
    )
    sheet = workbook.active

    assert sheet["C2"].value is None
    assert sheet.column_dimensions["C"].width == pytest.approx((18 * 9 - 5) / 7, abs=0.1)
    assert sheet.row_dimensions[1].height == pytest.approx(32 * 0.75, abs=0.1)
    assert sheet.row_dimensions[2].height == pytest.approx(120 * 0.75, abs=0.1)


def test_recognition_exception_sheet_is_created_even_when_empty() -> None:
    workbook = Workbook()
    sheet = workbook.active
    append_xlsx_rows(sheet, recognition_report_headers(), [["鞋款A", "黑色", "", "42", 1, "", "鞋款A 黑色 42"]])
    append_recognition_exception_sheet(workbook, [])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    loaded = load_workbook(buffer)

    assert RECOGNITION_EXCEPTION_SHEET_TITLE in loaded.sheetnames
    assert loaded[RECOGNITION_EXCEPTION_SHEET_TITLE].max_row == 1


def test_report_quantity_value_accepts_common_text_formats() -> None:
    assert report_quantity_value("*2") == 2
    assert report_quantity_value("2件") == 2
    assert report_quantity_value("1") == 1
    assert report_quantity_value("") == 1


def test_recognition_report_headers_are_business_columns() -> None:
    assert recognition_report_headers() == ["商品", "销售属性1", "图片", "销售属性2", "数量", "备注", "图片匹配文本"]
