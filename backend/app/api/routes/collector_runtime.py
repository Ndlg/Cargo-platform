from datetime import datetime, timedelta, timezone
import hashlib
from io import BytesIO
import json
from pathlib import Path
import re
from urllib.parse import quote
from typing import Annotated, Any
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, Body, Depends, Header, HTTPException, Query, status
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PillowImage, UnidentifiedImageError
from pydantic import BaseModel, Field, field_validator, model_validator
from sqlalchemy import or_, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.context import CurrentUser
from app.core.database import get_db
from app.core.deps import get_current_user, get_workspace_id, require_write
from app.core.security import create_collector_token, hash_collector_token
from app.models import (
    CaptureTask,
    Collector,
    ExportHeaderDefinition,
    ImageAsset,
    Product,
    ProductSku,
    RawCaptureRecord,
    StandardDetail,
    StandardDetailBatch,
    Workspace,
)
from app.repositories.base import model_to_dict
from app.api.routes.product_sku_linking import (
    ProductMatchingScope,
    ProductSkuLinkingPreviewRequest,
    preview_with_rules as preview_product_sku_with_rules,
    rows_for_preview as product_sku_rows_for_preview,
    saved_rule_payloads as saved_product_sku_rule_payloads,
)
from app.services.collection_contract import (
    build_raw_capture_record,
)
from app.services.collector_enrollment import build_connection_code, normalize_public_base_url
from app.services.order_row_reader import (
    order_row_sample_inputs_from_records,
    raw_records_for_task,
    task_waybill_counts,
)
from app.services.product_sku_linking import exportable_product_sku_linking_result
from app.services.regression_coverage import (
    analyze_waybill_coverage,
    recognition_exception_text,
    recognition_row_is_exportable,
)
from app.services.recognition_rule_packs import (
    RULE_PACK_MISSING_STATUS,
    active_recognition_rule_pack,
)
from app.services.waybill_reading import read_waybill_samples


router = APIRouter()

COLLECTOR_HEARTBEAT_TIMEOUT = timedelta(seconds=60)
COLLECTOR_ENROLLMENT_TIMEOUT = timedelta(minutes=10)
COLLECTOR_TASK_WINDOW_PROTOCOL = 2
COLLECTOR_TASK_WINDOW_LEASE = timedelta(seconds=30)
COLLECTOR_PROTOCOL_LEASE = timedelta(seconds=60)

COLLECTOR_CLIENT_ARCHIVE_ROOT = "Cargo Platform 采集器"
COLLECTOR_CLIENT_RELEASE_EXE = Path("dist") / "Cargo Platform 采集器.exe"
COLLECTOR_CLIENT_RELEASE_MANIFEST = Path("dist") / "collector-manifest.json"
COLLECTOR_CLIENT_MANIFEST_SCHEMA_VERSION = 1
RAW_CAPTURE_BATCH_MAX_RECORDS = 100
RAW_CAPTURE_PAYLOAD_MAX_CHARS = 2_000_000
RAW_CAPTURE_SOURCE_COLUMNS_MAX_CHARS = 20_000
BUSINESS_DOWNLOAD_TIMEZONE = timezone(timedelta(hours=8))
BUSINESS_REPORT_DOWNLOAD_PREFIX = "订单整理文档"
COLLECTOR_PENDING_MACHINE_NAME = "等待业务机上报机器名"
DEFAULT_COLLECTOR_DISPLAY_NAMES = {
    "",
    "Cargo Platform 采集器",
    "业务机采集器",
    "本机采集器",
    "采集器",
    COLLECTOR_PENDING_MACHINE_NAME,
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_optional_text(value: Any) -> str:
    return str(value or "").strip()


def is_default_collector_display_name(value: Any) -> bool:
    return clean_optional_text(value) in DEFAULT_COLLECTOR_DISPLAY_NAMES


def collector_display_name(
    value: Any,
    *,
    source_machine: Any = None,
    collector_id: Any = None,
) -> str:
    name = clean_optional_text(value)
    if not is_default_collector_display_name(name):
        return name
    machine = clean_optional_text(source_machine)
    if machine:
        return machine
    identity = clean_optional_text(collector_id)
    if identity and not identity.startswith("collector-"):
        return identity
    return COLLECTOR_PENDING_MACHINE_NAME


def parse_utc_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def parse_collector_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=BUSINESS_DOWNLOAD_TIMEZONE)
    return parsed.astimezone(timezone.utc)


def collector_record_is_inside_task_window(item: Any, task: CaptureTask) -> bool:
    captured_at = parse_collector_datetime(item.captured_at)
    if captured_at is None:
        return True
    started_at = parse_utc_datetime(task.started_at)
    ended_at = parse_utc_datetime(task.ended_at)
    return not (
        (started_at is not None and captured_at < started_at.replace(microsecond=0))
        or (ended_at is not None and captured_at > ended_at)
    )


def collector_heartbeat_is_stale(collector: Collector) -> bool:
    if collector.online_status != "online":
        return False
    last_heartbeat_at = parse_utc_datetime(collector.last_heartbeat_at)
    if last_heartbeat_at is None:
        return True
    return datetime.now(timezone.utc) - last_heartbeat_at > COLLECTOR_HEARTBEAT_TIMEOUT


def get_workspace_tenant_id(db: Session, workspace_id: int) -> int | None:
    workspace = db.get(Workspace, workspace_id)
    if workspace is None or workspace.is_deleted:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Workspace access denied.")
    return workspace.tenant_id


def collector_client_source_dir() -> Path:
    return Path(__file__).resolve().parents[4] / "collector-client"


def collector_client_archive_path(name: str) -> str:
    return str(Path(COLLECTOR_CLIENT_ARCHIVE_ROOT) / name)


def collector_client_file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validated_collector_client_release(source_dir: Path) -> tuple[Path, Path, dict[str, Any]]:
    exe_path = source_dir / COLLECTOR_CLIENT_RELEASE_EXE
    manifest_path = source_dir / COLLECTOR_CLIENT_RELEASE_MANIFEST
    if not exe_path.is_file() or not manifest_path.is_file():
        raise FileNotFoundError(
            "采集器发布包缺失，需要同时构建 collector-client/dist/Cargo Platform 采集器.exe "
            "和 collector-manifest.json。"
        )

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ValueError("采集器发布清单不是有效 JSON。") from exc
    if not isinstance(manifest, dict):
        raise ValueError("采集器发布清单格式无效。")

    required_text_fields = (
        "release_version",
        "client_version",
        "git_sha",
        "python_version",
        "pyinstaller_version",
        "sha256",
    )
    if manifest.get("schema_version") != COLLECTOR_CLIENT_MANIFEST_SCHEMA_VERSION:
        raise ValueError("采集器发布清单版本不受支持。")
    if manifest.get("artifact") != COLLECTOR_CLIENT_RELEASE_EXE.name:
        raise ValueError("采集器发布清单中的文件名不匹配。")
    if any(not isinstance(manifest.get(field), str) or not manifest[field].strip() for field in required_text_fields):
        raise ValueError("采集器发布清单缺少必填字段。")
    if manifest["release_version"] != get_settings().app_version:
        raise ValueError("采集器发布版本与平台版本不一致。")
    if not re.fullmatch(r"[0-9a-fA-F]{40}", manifest["git_sha"]):
        raise ValueError("采集器发布清单中的 Git SHA 无效。")
    if not re.fullmatch(r"[0-9a-fA-F]{64}", manifest["sha256"]):
        raise ValueError("采集器发布清单中的 SHA-256 无效。")

    file_size = exe_path.stat().st_size
    if not isinstance(manifest.get("size"), int) or manifest["size"] != file_size:
        raise ValueError("采集器 EXE 文件大小与发布清单不一致。")
    with exe_path.open("rb") as handle:
        if handle.read(2) != b"MZ":
            raise ValueError("采集器发布文件不是有效的 Windows EXE。")
    if collector_client_file_sha256(exe_path) != manifest["sha256"].lower():
        raise ValueError("采集器 EXE 的 SHA-256 与发布清单不一致。")

    return exe_path, manifest_path, manifest


def require_collector_client_release(source_dir: Path) -> tuple[Path, Path, dict[str, Any]]:
    try:
        return validated_collector_client_release(source_dir)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail=str(exc)) from exc


def write_collector_client_version(zip_file: ZipFile, manifest: dict[str, Any], *, mode: str) -> None:
    zip_file.writestr(
        collector_client_archive_path("VERSION.txt"),
        (
            f"version={manifest['client_version']}\n"
            f"release_version={manifest['release_version']}\n"
            f"git_sha={manifest['git_sha']}\n"
            f"sha256={manifest['sha256']}\n"
            f"mode={mode}\n"
            "package=single-exe-token-collector\n"
            "features=single-exe,no-console-window,token-only,no-password-on-business-machine,server-reconnect-wait,remote-disconnect-guard\n"
        ),
    )


def collector_client_parameter_guide() -> str:
    return (
        "Cargo Platform 采集器安装说明\n"
        "\n"
        "文件：Cargo Platform 采集器.exe\n"
        "\n"
        "1. 双击 Cargo Platform 采集器.exe。\n"
        "2. 在安装器中粘贴管理页面生成的 CP1 连接码。\n"
        "3. 等待安装器提示登记完成。\n"
        "\n"
        "连接码仅用于本次登记；不要在业务机输入或保存系统账号密码。\n"
    )


def write_collector_client_release(zip_file: ZipFile, source_dir: Path, *, mode: str) -> None:
    exe_path, manifest_path, manifest = require_collector_client_release(source_dir)
    write_collector_client_version(zip_file, manifest, mode=mode)
    zip_file.write(exe_path, collector_client_archive_path("Cargo Platform 采集器.exe"))
    zip_file.write(manifest_path, collector_client_archive_path("collector-manifest.json"))
    zip_file.writestr(
        collector_client_archive_path("参数说明.txt"),
        collector_client_parameter_guide(),
    )


def build_collector_client_archive(mode: str = "cli") -> BytesIO:
    source_dir = collector_client_source_dir()
    if not source_dir.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collector client package not found.")

    archive = BytesIO()
    with ZipFile(archive, "w", ZIP_DEFLATED) as zip_file:
        write_collector_client_release(zip_file, source_dir, mode=mode)
    archive.seek(0)
    return archive


def collector_client_release_status() -> dict[str, Any]:
    source_dir = collector_client_source_dir()
    base_status: dict[str, Any] = {
        "package_version": "",
        "release_version": "",
        "sha256": "",
        "release_available": False,
        "archive_name": "订单整理系统采集器.zip",
        "release_exe": str(COLLECTOR_CLIENT_RELEASE_EXE).replace("\\", "/"),
    }
    try:
        _, _, manifest = validated_collector_client_release(source_dir)
    except FileNotFoundError as exc:
        return {**base_status, "status": "missing", "message": str(exc)}
    except ValueError as exc:
        return {**base_status, "status": "invalid", "message": str(exc)}
    return {
        **base_status,
        "package_version": manifest["client_version"],
        "release_version": manifest["release_version"],
        "sha256": manifest["sha256"],
        "release_available": True,
        "status": "ready",
        "message": "采集器发布包已校验并就绪。",
    }


class CollectorRegisterRequest(BaseModel):
    collector_id: str | None = Field(default=None, max_length=128)
    collector_name: str = Field(default="", max_length=128)
    source_machine: str | None = Field(default=None, max_length=128)
    client_version: str | None = Field(default=None, max_length=64)
    remark: str | None = None
    public_base_url: str = Field(min_length=1, max_length=2048)

    @field_validator("public_base_url")
    @classmethod
    def public_base_url_must_be_http(cls, value: str) -> str:
        return normalize_public_base_url(value)


class CollectorEnrollmentRequest(BaseModel):
    token: str = Field(min_length=1, max_length=512)
    collector_id: str | None = Field(default=None, max_length=128)
    collector_name: str = Field(default="", max_length=128)
    source_machine: str | None = Field(default=None, max_length=128)
    client_version: str | None = Field(default=None, max_length=64)


class CollectorRepairCodeRequest(BaseModel):
    public_base_url: str = Field(min_length=1, max_length=2048)

    @field_validator("public_base_url")
    @classmethod
    def public_base_url_must_be_http(cls, value: str) -> str:
        return normalize_public_base_url(value)


class CaptureStartRequest(BaseModel):
    name: str | None = Field(default=None, max_length=128)
    collector_id: int | None = None


class CaptureStopRequest(BaseModel):
    task_id: int | None = None


class CollectorHeartbeatRequest(BaseModel):
    collector_id: str | None = Field(default=None, max_length=128)
    collector_name: str | None = Field(default=None, max_length=128)
    source_machine: str | None = Field(default=None, max_length=128)
    client_version: str | None = Field(default=None, max_length=64)
    runtime_status: str | None = Field(default=None, max_length=32)
    adapter_status: dict[str, Any] | None = None
    queue_size: int | None = None
    last_error: str | None = None
    last_upload_at: str | None = Field(default=None, max_length=64)
    last_reconnect_reason: str | None = Field(default=None, max_length=32)
    tracked_task_ids: list[int] = Field(default_factory=list, max_length=32)
    assignment_protocol_version: int = Field(default=1, ge=1, le=COLLECTOR_TASK_WINDOW_PROTOCOL)
    pending_captured_at: str | None = Field(default=None, max_length=64)
    pending_captured_until: str | None = Field(default=None, max_length=64)
    pending_row_count: int = Field(default=0, ge=0, le=100_000)


class RawCaptureRecordPayload(BaseModel):
    """Collector upload payload whose public persisted output is raw_capture_record."""

    document_id: str | None = Field(default=None, max_length=128)
    source_machine: str | None = Field(default=None, max_length=128)
    source_component: str | None = Field(default=None, max_length=128)
    source_index: str | None = Field(default=None, max_length=128)
    dedupe_key: str | None = Field(default=None, max_length=255)
    waybill_mode: str | None = Field(default=None, max_length=128)
    payload_format: str = Field(default="unknown", max_length=32)
    raw_payload: str = Field(min_length=1, max_length=RAW_CAPTURE_PAYLOAD_MAX_CHARS)
    source_columns: dict[str, Any] | None = None
    parsed_payload: dict[str, Any] | None = None
    captured_at: str | None = Field(default=None, max_length=64)

    @field_validator("source_columns")
    @classmethod
    def source_columns_must_be_audit_sized(
        cls,
        value: dict[str, Any] | None,
    ) -> dict[str, Any] | None:
        if value is None:
            return None
        serialized = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        if len(serialized) > RAW_CAPTURE_SOURCE_COLUMNS_MAX_CHARS:
            raise ValueError(
                f"source_columns must be at most {RAW_CAPTURE_SOURCE_COLUMNS_MAX_CHARS} JSON characters."
            )
        return value


class RawCaptureBatchRequest(BaseModel):
    task_id: int
    assignment_protocol_version: int = Field(default=1, ge=1, le=COLLECTOR_TASK_WINDOW_PROTOCOL)
    records: list[RawCaptureRecordPayload] = Field(
        min_length=1,
        max_length=RAW_CAPTURE_BATCH_MAX_RECORDS,
    )

    @model_validator(mode="after")
    def v2_records_require_source_identity(self) -> "RawCaptureBatchRequest":
        if self.assignment_protocol_version >= COLLECTOR_TASK_WINDOW_PROTOCOL and any(
            not record.source_component or not record.source_index
            for record in self.records
        ):
            raise ValueError("Protocol v2 records require source_component and source_index.")
        return self


class ParseRecordsRequest(BaseModel):
    task_id: int | None = None
    force: bool = False


class ArchiveCaptureDataRequest(BaseModel):
    days_before: int | None = Field(default=None, ge=0, le=3650)


class DeleteArchivedCaptureDataRequest(BaseModel):
    confirm_text: str
    days_before: int | None = Field(default=None, ge=0, le=3650)


def public_collector(collector: Collector) -> dict[str, Any]:
    data = model_to_dict(collector)
    if not collector_heartbeat_is_stale(collector):
        return data

    status_payload = data.get("status_payload")
    if isinstance(status_payload, str):
        try:
            status_payload = json.loads(status_payload)
        except json.JSONDecodeError:
            status_payload = {}
    elif isinstance(status_payload, dict):
        status_payload = dict(status_payload)
    else:
        status_payload = {}

    status_payload["runtime_status"] = "stale"
    status_payload["stale_reason"] = "heartbeat_timeout"
    status_payload["heartbeat_timeout_seconds"] = int(COLLECTOR_HEARTBEAT_TIMEOUT.total_seconds())
    data["online_status"] = "offline"
    data["status_payload"] = status_payload
    return data


def public_task(task: CaptureTask) -> dict[str, Any]:
    return model_to_dict(task)


def json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, indent=2)


def text_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def int_value(value: Any) -> int | None:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def recognition_summary(rows: list[dict[str, Any]]) -> dict[str, int]:
    summary = {
        "total": 0,
        "matched": 0,
        "product_unmatched": 0,
        "sku_unmatched": 0,
        "conflict": 0,
    }
    for row in rows:
        if not row.get("coverage_only"):
            summary["total"] += 1
        status_text = text_value(row.get("status"))
        if status_text in summary:
            summary[status_text] += 1
    return summary


def source_component_label(component: Any) -> str:
    component_text = text_value(component)
    if component_text == "cloud-print-client":
        return "抖店打印组件"
    if component_text == "cainiao-cnprint":
        return "菜鸟打印组件"
    return component_text or "-"


def raw_record_collector_label(record: RawCaptureRecord, collectors_by_id: dict[int, Collector]) -> str:
    collector = collectors_by_id.get(int(record.collector_id or 0))
    if collector is not None and text_value(collector.collector_name):
        return text_value(collector.collector_name)
    return text_value(record.source_machine) or "-"


def infer_size_text(*values: Any) -> str:
    text = "\n".join(text_value(value) for value in values if text_value(value))
    labeled = re.search(r"(?:鞋码|尺码|码数|尺碼)\s*[:：]?\s*([2-4]\d(?:\.5)?|50|[XSML]{1,4})", text, re.I)
    if labeled:
        return labeled.group(1)
    generic = re.search(r"(?<!\d)([2-4]\d(?:\.5)?|50)(?!\d)", text)
    return generic.group(1) if generic else ""


CUSTOM_FIELD_FALLBACKS = {
    "custom_spec_text": "custom_sales_attr1_text",
    "custom_size_text": "custom_sales_attr2_text",
    "quantity": "custom_quantity_text",
}


def custom_item_export_values(
    base_values: dict[str, Any],
    item: dict[str, Any],
    *,
    item_index: int,
    item_count: int,
) -> dict[str, Any]:
    values = dict(base_values)
    remark_text = text_value(item.get("remark_text"))
    sales_attr1 = text_value(item.get("sales_attr1_text") or item.get("spec_text"))
    sales_attr2 = text_value(item.get("sales_attr2_text") or item.get("size_text"))
    quantity_text = text_value(item.get("quantity_text"))
    values.update(
        {
            "custom_item_index": item_index,
            "custom_item_count": item_count,
            "custom_item_key": f"{base_values.get('raw_record_id') or base_values.get('raw_document_id')}-{item_index}",
            "custom_product_text": text_value(item.get("product_text")),
            "custom_sales_attr1_text": sales_attr1,
            "custom_sales_attr2_text": sales_attr2,
            "custom_spec_text": text_value(item.get("spec_text")) or sales_attr1,
            "custom_size_text": sales_attr2,
            "custom_quantity_text": quantity_text,
            "custom_item_remark_text": remark_text,
            "custom_item_raw_text": text_value(item.get("raw_text")),
        }
    )
    if quantity_text:
        values["quantity"] = quantity_text
    elif item_count > 1:
        values["quantity"] = ""
    return values


def standard_detail_export_rows(detail: StandardDetail) -> list[dict[str, Any]]:
    values = detail.field_values or {}
    custom_items = values.get("custom_items")
    if not isinstance(custom_items, list) or not custom_items:
        return [values]

    item_dicts = [item for item in custom_items if isinstance(item, dict)]
    if not item_dicts:
        return [values]

    item_count = len(item_dicts)
    return [
        custom_item_export_values(values, item, item_index=index, item_count=item_count)
        for index, item in enumerate(item_dicts, start=1)
    ]


def export_field_value(field_code: str, values: dict[str, Any]) -> Any:
    if field_code == "inferred_size":
        return infer_size_text(
            values.get("custom_sales_attr2_text"),
            values.get("custom_size_text"),
            values.get("custom_item_remark_text"),
            values.get("spec_text"),
            values.get("product_short_text"),
            values.get("product_full_text"),
            values.get("custom_area_raw_text"),
        )
    if field_code == "product_display_text":
        is_woda_custom_row = values.get("source_platform") == "woda" or values.get("custom_area_raw_text") not in (None, "")
        if not is_woda_custom_row:
            return (
                values.get("product_short_text")
                or values.get("product_full_text")
                or values.get("custom_item_raw_text")
                or values.get("custom_product_text")
                or values.get("custom_area_raw_text")
                or ""
            )
        return (
            values.get("custom_product_text")
            or values.get("product_short_text")
            or values.get("product_full_text")
            or values.get("custom_area_raw_text")
            or ""
        )
    value = values.get(field_code)
    if value in (None, "") and field_code in CUSTOM_FIELD_FALLBACKS:
        return values.get(CUSTOM_FIELD_FALLBACKS[field_code], "")
    return value


RECOGNITION_REPORT_HEADERS = ["商品", "销售属性1", "图片", "销售属性2", "数量", "备注", "图片匹配文本"]

RECOGNITION_REPORT_FIELD_DEFINITIONS: dict[str, dict[str, Any]] = {
    "product_name": {"label": "商品", "width": 16},
    "sales_attr1": {"label": "销售属性1", "width": 24},
    "sku_image": {"label": "图片", "width": 18},
    "sales_attr2": {"label": "销售属性2", "width": 18},
    "quantity": {"label": "数量", "width": 12},
    "remark": {"label": "备注", "width": 18},
    "image_match_text": {"label": "图片匹配文本", "width": 42},
}

RECOGNITION_REPORT_DEFAULT_FIELD_ORDER = [
    "product_name",
    "sales_attr1",
    "sku_image",
    "sales_attr2",
    "quantity",
    "remark",
    "image_match_text",
]

RECOGNITION_REPORT_OUTPUT_MODES = {"merged_sheet", "stall_sheet", "stall_workbooks"}
DEFAULT_RECOGNITION_REPORT_OUTPUT_MODE = "stall_sheet"

RECOGNITION_EXCEPTION_HEADERS = ["图片匹配文本"]
RECOGNITION_EXCEPTION_SHEET_TITLE = "异常面单"
EXPORT_PRODUCT_SKU_LINKING_CONTRACT = "product-sku-linking-results-v1"
EXPORT_PRODUCT_SKU_LINKING_RESULTS_KEY = "product_sku_linking_results"
EXPORT_PRODUCT_SKU_LINKING_RESULT_KEY = "product_sku_linking_result"
EXPORT_PRODUCT_SKU_LINKING_PENDING_STATUS = "pending"

RECOGNITION_REPORT_LEGACY_LABELS = {
    "product_name": {"商品名称"},
    "sku_image": {"SKU图片"},
}

REPORT_IMAGE_SIZE = 88
REPORT_ROW_HEIGHT = 86
REPORT_HEADER_ROW_HEIGHT = 26
REPORT_COLUMN_WIDTH_PIXEL_RATIO = 9
EXCEL_COLUMN_PIXEL_PADDING = 5
EXCEL_COLUMN_UNIT_PIXELS = 7
EXCEL_POINTS_PER_PIXEL = 0.75


def bounded_int(value: Any, default: int, min_value: int, max_value: int) -> int:
    try:
        parsed = int(round(float(value)))
    except (TypeError, ValueError):
        return default
    return min(max(parsed, min_value), max_value)


def report_layout_width_to_excel_width(value: Any) -> float:
    layout_width = bounded_int(value, 12, 8, 60)
    preview_pixels = layout_width * REPORT_COLUMN_WIDTH_PIXEL_RATIO
    return round(max(8, (preview_pixels - EXCEL_COLUMN_PIXEL_PADDING) / EXCEL_COLUMN_UNIT_PIXELS), 2)


def report_layout_height_to_excel_points(value: Any) -> float:
    height_pixels = bounded_int(value, REPORT_ROW_HEIGHT, 18, 220)
    return round(height_pixels * EXCEL_POINTS_PER_PIXEL, 2)


def default_recognition_report_layout() -> dict[str, Any]:
    return {
        "columns": [
            {
                "key": key,
                "label": RECOGNITION_REPORT_FIELD_DEFINITIONS[key]["label"],
                "visible": True,
                "width": RECOGNITION_REPORT_FIELD_DEFINITIONS[key]["width"],
            }
            for key in RECOGNITION_REPORT_DEFAULT_FIELD_ORDER
        ],
        "header_row_height": REPORT_HEADER_ROW_HEIGHT,
        "row_height": REPORT_ROW_HEIGHT,
        "image_width": REPORT_IMAGE_SIZE,
        "image_height": REPORT_IMAGE_SIZE,
        "image_offset_x": 0,
        "image_offset_y": 0,
        "stack_sales_attr1": False,
        "stack_sales_attr2": False,
        "output_mode": DEFAULT_RECOGNITION_REPORT_OUTPUT_MODE,
    }


def normalize_recognition_report_layout(raw_layout: Any | None = None) -> dict[str, Any]:
    default_layout = default_recognition_report_layout()
    payload = raw_layout if isinstance(raw_layout, dict) else {}
    source_columns = payload.get("columns")
    if not isinstance(source_columns, list):
        source_columns = []

    columns: list[dict[str, Any]] = []
    used_keys: set[str] = set()
    for source_column in source_columns:
        if not isinstance(source_column, dict):
            continue
        key = str(source_column.get("key") or "")
        definition = RECOGNITION_REPORT_FIELD_DEFINITIONS.get(key)
        if definition is None or key in used_keys:
            continue
        used_keys.add(key)
        label = str(source_column.get("label") or definition["label"]).strip() or definition["label"]
        if label in RECOGNITION_REPORT_LEGACY_LABELS.get(key, set()):
            label = definition["label"]
        columns.append(
            {
                "key": key,
                "label": label[:40],
                "visible": (
                    True
                    if key in RECOGNITION_REPORT_DEFAULT_FIELD_ORDER
                    else source_column.get("visible") is not False
                ),
                "width": bounded_int(source_column.get("width"), int(definition["width"]), 8, 60),
            }
        )

    for key in RECOGNITION_REPORT_DEFAULT_FIELD_ORDER:
        if key in used_keys:
            continue
        definition = RECOGNITION_REPORT_FIELD_DEFINITIONS[key]
        columns.append(
            {
                "key": key,
                "label": definition["label"],
                "visible": True,
                "width": definition["width"],
            }
        )

    if not any(column["visible"] for column in columns):
        for column in columns:
            column["visible"] = True

    output_mode = str(payload.get("output_mode", payload.get("outputMode")) or default_layout["output_mode"])
    if output_mode not in RECOGNITION_REPORT_OUTPUT_MODES:
        output_mode = str(default_layout["output_mode"])

    return {
        "columns": columns,
        "header_row_height": bounded_int(
            payload.get("header_row_height", payload.get("headerRowHeight")),
            int(default_layout["header_row_height"]),
            18,
            80,
        ),
        "row_height": bounded_int(
            payload.get("row_height", payload.get("rowHeight")),
            int(default_layout["row_height"]),
            24,
            180,
        ),
        "image_width": bounded_int(
            payload.get("image_width", payload.get("imageWidth")),
            int(default_layout["image_width"]),
            32,
            220,
        ),
        "image_height": bounded_int(
            payload.get("image_height", payload.get("imageHeight")),
            int(default_layout["image_height"]),
            32,
            220,
        ),
        "image_offset_x": bounded_int(
            payload.get("image_offset_x", payload.get("imageOffsetX")),
            int(default_layout["image_offset_x"]),
            0,
            220,
        ),
        "image_offset_y": bounded_int(
            payload.get("image_offset_y", payload.get("imageOffsetY")),
            int(default_layout["image_offset_y"]),
            0,
            220,
        ),
        "stack_sales_attr1": bool(payload.get("stack_sales_attr1", payload.get("stackSalesAttr1", False))),
        "stack_sales_attr2": bool(payload.get("stack_sales_attr2", payload.get("stackSalesAttr2", False))),
        "output_mode": output_mode,
    }


def recognition_report_layout_from_query(layout: str | None) -> dict[str, Any]:
    if not layout:
        return normalize_recognition_report_layout()
    try:
        parsed = json.loads(layout)
    except json.JSONDecodeError:
        return normalize_recognition_report_layout()
    return normalize_recognition_report_layout(parsed)


def visible_recognition_report_columns(layout: dict[str, Any]) -> list[dict[str, Any]]:
    return [column for column in layout["columns"] if column.get("visible") is not False]


def recognition_status_label(status_text: str) -> str:
    return {
        "matched": "已匹配",
        "product_unmatched": "商品未命中",
        "sku_unmatched": "SKU未命中",
        "conflict": "冲突",
    }.get(status_text, status_text or "-")


def recognition_image_label(row: dict[str, Any]) -> str:
    return text_value(row.get("image_label"))


def recognition_stall_name(row: dict[str, Any]) -> str:
    return text_value(row.get("stall_name")) or "未设置档口"


def report_quantity_value(value: Any, *, default: int = 1) -> int:
    text = text_value(value)
    if not text:
        return default
    compact = re.sub(r"\s+", "", text)
    match = re.fullmatch(r"[*xX×]?(\d+)(?:件|个|個|双|雙|条|條|套|只|瓶|包|箱)?", compact)
    if not match:
        return default
    parsed = int(match.group(1))
    return parsed if parsed > 0 else default


def report_quantity_default(row: dict[str, Any]) -> int:
    if (int_value(row.get("item_count")) or 0) > 1 and not text_value(row.get("quantity_text")):
        return 0
    return 1


def report_spec_text(row: dict[str, Any]) -> str:
    return text_value(row.get("sales_attr1_text")) or "-"


def report_sales_attr2_values(value: Any) -> list[str]:
    text = text_value(value)
    if not text:
        return ["-"]
    return [text]


def natural_report_sort_key(value: Any) -> tuple[int, float | str, str]:
    text = text_value(value) or "-"
    match = re.search(r"\d+(?:\.\d+)?", text)
    if match:
        return (0, float(match.group(0)), text.lower())
    return (1, text.lower(), text.lower())


def sorted_report_values(values: list[str]) -> list[str]:
    return sorted([value for value in values if value], key=natural_report_sort_key)


def unique_joined_report_values(values: list[Any], separator: str = "\n") -> str:
    unique: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = text_value(value)
        if not text or text in seen:
            continue
        seen.add(text)
        unique.append(text)
    return separator.join(unique)


def expanded_sales_attr2_values(row: dict[str, Any]) -> list[str]:
    tokens = report_sales_attr2_values(row.get("sales_attr2_text"))
    quantity = report_quantity_value(row.get("quantity_text"), default=report_quantity_default(row))
    if len(tokens) > 1:
        return tokens
    return [tokens[0] or "-"] * quantity


def recognition_report_row_is_exportable(row: dict[str, Any]) -> bool:
    return recognition_row_is_exportable(row)


def recognition_report_base_line_item(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "product_category": text_value(row.get("product_name")) or "-",
        "product_id": int_value(row.get("product_id")),
        "candidate_key": text_value(row.get("candidate_key")),
        "stall_id": int_value(row.get("stall_id")),
        "stall_name": recognition_stall_name(row),
        "spec": report_spec_text(row),
        "image_label": recognition_image_label(row),
        "sku_id": int_value(row.get("sku_id")),
        "sku_image_asset_id": int_value(row.get("sku_image_asset_id")),
        "size_text": text_value(row.get("sales_attr2_text")) or "-",
        "quantity": report_quantity_value(row.get("quantity_text"), default=report_quantity_default(row)),
        "remark_text": text_value(row.get("remark_text")),
        "image_match_text": text_value(row.get("image_match_text")),
    }


def recognition_report_line_items(
    rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    normalized_layout = normalize_recognition_report_layout(layout)
    report_rows: list[dict[str, Any]] = []
    for row in rows:
        if not recognition_report_row_is_exportable(row):
            continue
        report_rows.append(recognition_report_base_line_item(row))

    if not normalized_layout["stack_sales_attr1"]:
        return report_rows

    grouped: dict[str, dict[str, Any]] = {}
    for row in report_rows:
        key = ":".join(
            [
                text_value(row.get("stall_id")) or text_value(row.get("stall_name")),
                text_value(row.get("product_id")) or text_value(row.get("product_category")),
                text_value(row.get("sku_id")) or text_value(row.get("spec")),
                text_value(row.get("sku_image_asset_id")) or "0",
                "grouped",
            ]
        )
        group = grouped.setdefault(
            key,
            {
                **row,
                "spec_values": [],
                "size_values": [],
                "remark_values": [],
                "image_match_text_values": [],
                "quantity": 0,
            },
        )
        group["spec_values"].append(text_value(row.get("spec")))
        group["size_values"].extend(
            expanded_sales_attr2_values(
                {
                    "sales_attr2_text": row.get("size_text"),
                    "quantity_text": row.get("quantity"),
                    "item_count": 1,
                }
            )
        )
        group["remark_values"].append(row.get("remark_text"))
        group["image_match_text_values"].append(row.get("image_match_text"))
        group["quantity"] += int_value(row.get("quantity")) or 0

    merged_rows: list[dict[str, Any]] = []
    for group in grouped.values():
        spec_values = list(group.pop("spec_values", []))
        size_values = list(group.pop("size_values", []))
        remark_values = list(group.pop("remark_values", []))
        image_match_text_values = list(group.pop("image_match_text_values", []))
        group["spec"] = " ".join(sorted_report_values(list(dict.fromkeys(spec_values)))) or "-"
        group["size_text"] = (
            " ".join(sorted_report_values(list(dict.fromkeys(size_values))))
            if normalized_layout["stack_sales_attr2"]
            else " ".join(sorted_report_values(size_values))
        ) or "-"
        group["remark_text"] = unique_joined_report_values(remark_values)
        group["image_match_text"] = unique_joined_report_values(image_match_text_values)
        merged_rows.append(group)

    return sorted(
        merged_rows,
        key=lambda row: (
            natural_report_sort_key(row.get("stall_name")),
            natural_report_sort_key(row.get("product_category")),
            natural_report_sort_key(row.get("spec")),
            natural_report_sort_key(row.get("size_text")),
        ),
    )


def recognition_report_cell_value(row: dict[str, Any], field_key: str) -> Any:
    if field_key == "product_name":
        return row["product_category"]
    if field_key == "sales_attr1":
        return row["spec"]
    if field_key == "sku_image":
        return ""
    if field_key == "sales_attr2":
        return row["size_text"]
    if field_key == "quantity":
        return row["quantity"]
    if field_key == "remark":
        return row.get("remark_text", "")
    if field_key == "image_match_text":
        return row.get("image_match_text", "")
    return ""


def recognition_report_headers(layout: dict[str, Any] | None = None) -> list[str]:
    normalized_layout = normalize_recognition_report_layout(layout)
    return [str(column["label"]) for column in visible_recognition_report_columns(normalized_layout)]


def recognition_report_export_rows(
    rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
) -> list[list[Any]]:
    normalized_layout = normalize_recognition_report_layout(layout)
    columns = visible_recognition_report_columns(normalized_layout)
    return recognition_report_export_rows_from_line_items(
        recognition_report_line_items(rows, normalized_layout),
        normalized_layout,
    )


def recognition_report_export_rows_from_line_items(
    report_rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
) -> list[list[Any]]:
    normalized_layout = normalize_recognition_report_layout(layout)
    columns = visible_recognition_report_columns(normalized_layout)
    return [
        [
            recognition_report_cell_value(row, str(column["key"]))
            for column in columns
        ]
        for row in report_rows
    ]


def recognition_report_rows_by_stall(report_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in report_rows:
        grouped.setdefault(recognition_stall_name(row), []).append(row)
    if grouped:
        return grouped
    return {"未设置档口": []}


def recognition_exception_export_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [recognition_exception_text(row)]
        for row in rows
        if not recognition_report_row_is_exportable(row)
    ]


def recognition_report_image_path(image: ImageAsset) -> Path | None:
    storage_root = Path(get_settings().storage_root).resolve()
    image_path = Path(image.file_path).resolve()
    if not image_path.is_relative_to(storage_root) or not image_path.is_file():
        return None
    return image_path


def recognition_report_image_buffer(
    image_path: Path,
    *,
    image_width: int = REPORT_IMAGE_SIZE,
    image_height: int = REPORT_IMAGE_SIZE,
) -> BytesIO | None:
    try:
        with PillowImage.open(image_path) as source:
            source.thumbnail((image_width, image_height))
            converted = source.convert("RGB")
            buffer = BytesIO()
            converted.save(buffer, format="PNG")
            buffer.seek(0)
            return buffer
    except (OSError, UnidentifiedImageError):
        return None


def attach_recognition_report_images(
    sheet,
    rows: list[dict[str, Any]],
    images_by_id: dict[int, ImageAsset],
    image_buffers: list[BytesIO],
    layout: dict[str, Any] | None = None,
) -> None:
    if not rows:
        return
    normalized_layout = normalize_recognition_report_layout(layout)
    image_column_index = next(
        (
            index
            for index, column in enumerate(visible_recognition_report_columns(normalized_layout), start=1)
            if column["key"] == "sku_image"
        ),
        None,
    )
    if image_column_index is None:
        return
    image_width = int(normalized_layout["image_width"])
    image_height = int(normalized_layout["image_height"])
    image_offset_x = int(normalized_layout["image_offset_x"])
    image_offset_y = int(normalized_layout["image_offset_y"])
    for row_number, row in enumerate(rows, start=2):
        image_asset_id = int_value(row.get("sku_image_asset_id"))
        if image_asset_id is None:
            continue
        image = images_by_id.get(image_asset_id)
        if image is None:
            continue
        image_path = recognition_report_image_path(image)
        if image_path is None:
            continue
        buffer = recognition_report_image_buffer(
            image_path,
            image_width=image_width,
            image_height=image_height,
        )
        if buffer is None:
            continue
        image_buffers.append(buffer)
        worksheet_image = WorksheetImage(buffer)
        worksheet_image.width = image_width
        worksheet_image.height = image_height
        worksheet_image.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=image_column_index - 1,
                colOff=pixels_to_EMU(image_offset_x),
                row=row_number - 1,
                rowOff=pixels_to_EMU(image_offset_y),
            ),
            ext=XDRPositiveSize2D(
                cx=pixels_to_EMU(image_width),
                cy=pixels_to_EMU(image_height),
            ),
        )
        sheet.add_image(worksheet_image)


def recognition_report_image_assets(
    db: Session,
    *,
    workspace_id: int,
    rows: list[dict[str, Any]],
) -> dict[int, ImageAsset]:
    image_asset_ids = sorted(
        {
            image_asset_id
            for row in rows
            if (image_asset_id := int_value(row.get("sku_image_asset_id"))) is not None
        }
    )
    if not image_asset_ids:
        return {}
    return {
        image.id: image
        for image in db.scalars(
            select(ImageAsset).where(
                ImageAsset.workspace_id == workspace_id,
                ImageAsset.id.in_(image_asset_ids),
                ImageAsset.is_deleted.is_(False),
            )
        ).all()
    }


def product_sku_linking_result_payloads(detail: StandardDetail) -> list[dict[str, Any]]:
    values = detail.field_values or {}
    results = values.get(EXPORT_PRODUCT_SKU_LINKING_RESULTS_KEY)
    if isinstance(results, list):
        return [item for item in results if isinstance(item, dict)]

    result = values.get(EXPORT_PRODUCT_SKU_LINKING_RESULT_KEY)
    if isinstance(result, dict):
        return [result]

    return []


def export_standard_fields_from_result(result: dict[str, Any]) -> dict[str, Any]:
    standard_fields = result.get("standard_fields")
    if isinstance(standard_fields, dict):
        return standard_fields
    return {}


def export_result_value(
    result: dict[str, Any],
    standard_fields: dict[str, Any],
    key: str,
    *fallback_keys: str,
) -> Any:
    for source_key in (key, *fallback_keys):
        value = result.get(source_key)
        if value not in (None, ""):
            return value
    return standard_fields.get(key, "")


def business_waybill_source_label(
    detail_number: int,
    *,
    item_index: int = 1,
    item_count: int = 1,
) -> str:
    parent_label = f"面单 {detail_number}"
    if item_count > 1:
        return f"{parent_label}-子项 {item_index}"
    return parent_label


def product_sku_linking_export_row(
    payload: dict[str, Any],
    *,
    source_identifiers: dict[str, Any],
    candidate_key_fallback: str,
    detail_number: int,
    item_index: int,
    item_count: int,
) -> dict[str, Any]:
    standard_fields = export_standard_fields_from_result(payload)
    status_text = text_value(payload.get("match_status")) or text_value(payload.get("status")) or "pending"
    image_value = payload.get("image")
    image_asset_id = int_value(payload.get("image_asset_id")) or int_value(payload.get("sku_image_asset_id"))
    image_label = text_value(payload.get("image_label"))
    if isinstance(image_value, dict):
        image_asset_id = image_asset_id or int_value(image_value.get("id"))
        image_label = image_label or text_value(image_value.get("name"))
    else:
        image_label = image_label or text_value(image_value)
    product_name = text_value(payload.get("product")) or text_value(payload.get("product_name"))
    sku_name = text_value(payload.get("sku")) or text_value(payload.get("sku_name"))
    sales_attr1 = text_value(export_result_value(payload, standard_fields, "sales_attr1", "sales_attr1_text"))
    sales_attr2 = text_value(export_result_value(payload, standard_fields, "sales_attr2", "sales_attr2_text"))
    quantity = text_value(export_result_value(payload, standard_fields, "quantity", "quantity_text"))
    remark = text_value(export_result_value(payload, standard_fields, "remark", "remark_text"))
    image_match_text = (
        text_value(payload.get("image_match_text"))
        or text_value(payload.get("match_text"))
    )
    stall_payload = payload.get("stall") if isinstance(payload.get("stall"), dict) else {}
    matched_rule = payload.get("matched_rule") if isinstance(payload.get("matched_rule"), dict) else {}
    stall_id = int_value(payload.get("stall_id")) or int_value(stall_payload.get("id"))
    stall_name = text_value(payload.get("stall_name")) or text_value(stall_payload.get("name"))

    return {
        "contract": EXPORT_PRODUCT_SKU_LINKING_CONTRACT,
        **source_identifiers,
        "candidate_key": text_value(payload.get("candidate_key")) or candidate_key_fallback,
        "source_label": business_waybill_source_label(
            detail_number,
            item_index=item_index,
            item_count=item_count,
        ),
        "item_index": item_index,
        "item_count": item_count,
        "product_text": text_value(standard_fields.get("product")),
        "sales_attr1_text": sales_attr1,
        "sales_attr2_text": sales_attr2,
        "quantity_text": quantity,
        "remark_text": remark,
        "image_match_text": image_match_text,
        "product_name": product_name,
        "product_id": int_value(payload.get("product_id")),
        "rule_id": int_value(matched_rule.get("id")),
        "stall_id": stall_id,
        "stall_name": stall_name,
        "sku_id": int_value(payload.get("sku_id")),
        "sku_name": sku_name,
        "sku_image_asset_id": image_asset_id,
        "image_label": image_label,
        "status": status_text,
        "reason": text_value(payload.get("exception_reason")) or text_value(payload.get("reason")),
        "match_type": "product_sku_linking_result",
        "match_field": "",
        "match_keyword": "",
    }


def product_sku_linking_result_row(
    detail: StandardDetail,
    result: dict[str, Any],
    *,
    detail_number: int,
    item_index: int,
    item_count: int,
) -> dict[str, Any]:
    return product_sku_linking_export_row(
        result,
        source_identifiers={"detail_id": detail.id},
        candidate_key_fallback=f"{detail.id}:{item_index}",
        detail_number=detail_number,
        item_index=item_index,
        item_count=item_count,
    )


def pending_product_sku_linking_row(detail: StandardDetail, *, detail_number: int) -> dict[str, Any]:
    message = "等待 Product/SKU Linking 模块输出后才能生成报货表。"
    return {
        "contract": EXPORT_PRODUCT_SKU_LINKING_CONTRACT,
        "detail_id": detail.id,
        "candidate_key": f"{detail.id}:pending-product-sku-linking",
        "source_label": f"面单 {detail_number}",
        "item_index": 1,
        "item_count": 1,
        "product_text": "",
        "sales_attr1_text": "",
        "sales_attr2_text": "",
        "quantity_text": "",
        "remark_text": "",
        "image_match_text": f"面单 {detail_number}：{message}",
        "product_name": "",
        "product_id": None,
        "sku_id": None,
        "sku_name": "",
        "sku_image_asset_id": None,
        "image_label": "",
        "status": EXPORT_PRODUCT_SKU_LINKING_PENDING_STATUS,
        "reason": message,
        "match_type": "product_sku_linking_result",
        "match_field": "",
        "match_keyword": "",
    }


def recognition_rows_from_product_sku_linking_results(details: list[StandardDetail]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for detail_number, detail in enumerate(details, start=1):
        payloads = product_sku_linking_result_payloads(detail)
        if not payloads:
            rows.append(pending_product_sku_linking_row(detail, detail_number=detail_number))
            continue
        item_count = len(payloads)
        for item_index, result in enumerate(payloads, start=1):
            rows.append(
                product_sku_linking_result_row(
                    detail,
                    result,
                    detail_number=detail_number,
                    item_index=item_index,
                    item_count=item_count,
                )
            )
    return rows


def pending_unmapped_waybill_product_sku_linking_row(
    sample: dict[str, Any],
    *,
    detail_number: int,
) -> dict[str, Any]:
    sample_text = text_value(sample.get("sample_text"))
    assignment = text_value(sample.get("capture_assignment"))
    message = {
        "timestamp_invalid_fallback": "这条打印记录的采集时间无效，已保留并隔离，请检查采集源时间。",
        "source_history_ambiguous": "打印数据库历史发生变化，这条记录已保留并隔离，请检查采集源。",
    }.get(assignment, "这张面单还没有生成五字段结果，无法进入商品匹配。")
    source_label = business_waybill_source_label(detail_number)
    return {
        "contract": EXPORT_PRODUCT_SKU_LINKING_CONTRACT,
        "detail_id": None,
        "raw_record_id": int_value(sample.get("raw_record_id")),
        "sample_id": text_value(sample.get("sample_id")),
        "candidate_key": f"{text_value(sample.get('sample_id')) or detail_number}:pending-order-row",
        "source_label": source_label,
        "item_index": 1,
        "item_count": 1,
        "product_text": "",
        "sales_attr1_text": "",
        "sales_attr2_text": "",
        "quantity_text": "",
        "remark_text": "",
        "image_match_text": sample_text or message,
        "product_name": "",
        "product_id": None,
        "sku_id": None,
        "sku_name": "",
        "sku_image_asset_id": None,
        "image_label": "",
        "status": EXPORT_PRODUCT_SKU_LINKING_PENDING_STATUS,
        "exception_code": assignment or None,
        "reason": message,
        "match_type": "product_sku_linking_result",
        "match_field": "",
        "match_keyword": "",
        "coverage_only": True,
    }


def unmapped_waybill_samples_for_task(
    db: Session,
    *,
    workspace_id: int,
    task_id: int,
    mapped_parent_sequences: set[int],
    mapped_raw_record_ids: set[int],
) -> list[dict[str, Any]]:
    raw_records = db.scalars(
        select(RawCaptureRecord)
        .where(
            RawCaptureRecord.workspace_id == workspace_id,
            RawCaptureRecord.task_id == task_id,
            RawCaptureRecord.is_deleted.is_(False),
            RawCaptureRecord.archived_at.is_(None),
        )
        .order_by(RawCaptureRecord.id.asc())
    ).all()

    unmapped_samples: list[dict[str, Any]] = []
    waybill_number = 0
    for raw_record in raw_records:
        samples = read_waybill_samples(raw_record)
        if not samples:
            samples = [{
                "sample_id": f"raw-{raw_record.id}-sample-1",
                "raw_record_id": raw_record.id,
                "sample_text": "",
            }]
        for sample in samples:
            waybill_number += 1
            if waybill_number in mapped_parent_sequences:
                continue
            if len(samples) == 1 and int(raw_record.id) in mapped_raw_record_ids:
                continue
            source_columns = (
                raw_record.source_columns
                if isinstance(raw_record.source_columns, dict)
                else {}
            )
            unmapped_samples.append(
                {
                    **sample,
                    "task_waybill_number": waybill_number,
                    "capture_assignment": source_columns.get("capture_assignment"),
                }
            )
    return unmapped_samples


def export_recognition_summary(rows: list[dict[str, Any]]) -> dict[str, int]:
    summary = recognition_summary(rows)
    for row in rows:
        status_text = text_value(row.get("status"))
        if status_text and status_text not in summary:
            summary[status_text] = 0
        if status_text and status_text not in {
            "total",
            "matched",
            "product_unmatched",
            "sku_unmatched",
            "conflict",
        }:
            summary[status_text] += 1
    return summary


CHILD_SOURCE_LABEL_SUFFIX_PATTERN = re.compile(r"-子(?:项\s*)?\d+$")
PARENT_SEQUENCE_PATTERN = re.compile(r"(?:第\d+批-第|面单\s*)(\d+)(?:单)?")


def recognition_waybill_count(rows: list[dict[str, Any]]) -> int:
    parent_labels: set[str] = set()
    for row in rows:
        source_label = text_value(row.get("source_label"))
        if not source_label:
            continue
        parent_label = CHILD_SOURCE_LABEL_SUFFIX_PATTERN.sub("", source_label)
        if parent_label:
            parent_labels.add(parent_label)
    return len(parent_labels) or len(rows)


def recognition_parent_sequences(rows: list[dict[str, Any]]) -> set[int]:
    sequences: set[int] = set()
    for row in rows:
        match = PARENT_SEQUENCE_PATTERN.search(text_value(row.get("source_label")))
        if match:
            sequences.add(int(match.group(1)))
    return sequences


def recognition_row_from_product_matching_preview(
    row: dict[str, Any],
    source: dict[str, Any],
    *,
    fallback_number: int,
) -> dict[str, Any]:
    payload = exportable_product_sku_linking_result(row)
    item_index = int_value(source.get("item_index")) or int_value(source.get("child_index")) or 1
    item_count = int_value(source.get("item_count")) or int_value(source.get("child_count")) or 1
    child_label = text_value(source.get("child_label"))
    raw_record_id = int_value(source.get("raw_record_id"))
    standard_detail = source.get("standard_detail")
    detail_id = standard_detail.id if isinstance(standard_detail, StandardDetail) else None
    result = product_sku_linking_export_row(
        payload,
        source_identifiers={
            "detail_id": detail_id,
            "raw_record_id": raw_record_id,
            "sample_id": child_label,
        },
        candidate_key_fallback=child_label or f"order-row:{fallback_number}",
        detail_number=fallback_number,
        item_index=item_index,
        item_count=item_count,
    )
    if child_label:
        parent_match = PARENT_SEQUENCE_PATTERN.search(child_label)
        parent_sequence = int(parent_match.group(1)) if parent_match else fallback_number
        result["source_label"] = business_waybill_source_label(
            parent_sequence,
            item_index=item_index,
            item_count=item_count,
        )
    source_row = source.get("row")
    source_status = text_value(getattr(source_row, "status", ""))
    if source_status == "special":
        result["status"] = "special"
        result["reason"] = text_value(getattr(source_row, "review_reason", "")) or "特殊面单，不进入商品/SKU/图片匹配。"
    return result


def recognition_rows_from_current_order_rows(
    db: Session,
    *,
    workspace_id: int,
    task_id: int,
) -> list[dict[str, Any]]:
    scope = ProductMatchingScope(
        scope_type="current_batch",
        task_id=task_id,
    )
    rows, sources = product_sku_rows_for_preview(
        db,
        workspace_id=workspace_id,
        payload=ProductSkuLinkingPreviewRequest(scope=scope),
    )
    rules = saved_product_sku_rule_payloads(db, workspace_id=workspace_id)
    preview = preview_product_sku_with_rules(db, workspace_id=workspace_id, rows=rows, rules=rules)
    return [
        recognition_row_from_product_matching_preview(row, source, fallback_number=index)
        for index, (row, source) in enumerate(zip(preview["rows"], sources, strict=False), start=1)
    ]


def recognition_rows_for_task(db: Session, *, workspace_id: int, task_id: int) -> list[dict[str, Any]]:
    rows = recognition_rows_from_current_order_rows(db, workspace_id=workspace_id, task_id=task_id)
    mapped_raw_record_ids = {
        raw_record_id
        for row in rows
        if (raw_record_id := int_value(row.get("raw_record_id"))) is not None
    }
    unmapped_samples = unmapped_waybill_samples_for_task(
        db,
        workspace_id=workspace_id,
        task_id=task_id,
        mapped_parent_sequences=recognition_parent_sequences(rows),
        mapped_raw_record_ids=mapped_raw_record_ids,
    )
    rows.extend(
        pending_unmapped_waybill_product_sku_linking_row(
            sample,
            detail_number=int_value(sample.get("task_waybill_number")) or len(rows) + 1,
        )
        for sample in unmapped_samples
    )
    return rows


def recognition_expected_waybill_count(db: Session, *, workspace_id: int, task_id: int) -> int:
    raw_record_count, waybill_count = task_waybill_counts(
        db,
        workspace_id=workspace_id,
        task_id=task_id,
    )
    if raw_record_count:
        return waybill_count
    return len(standard_details_for_task(db, workspace_id=workspace_id, task_id=task_id))


def require_complete_recognition_coverage(
    db: Session,
    *,
    workspace_id: int,
    task_id: int,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    expected = recognition_expected_waybill_count(db, workspace_id=workspace_id, task_id=task_id)
    covered = recognition_waybill_count(rows)
    if expected != covered:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="采集面单覆盖不完整，已停止生成报货文件。请先处理缺失面单。",
        )

    records = raw_records_for_task(db, workspace_id=workspace_id, task_id=task_id)
    if records:
        expected_parent_documents = [
            {
                "raw_record_id": int(sample["raw_record_id"]),
                "parent_sequence": int(sample["parent_sequence"]),
            }
            for sample in order_row_sample_inputs_from_records(records)
        ]
    else:
        expected_parent_documents = []
        for parent_sequence, detail in enumerate(
            standard_details_for_task(db, workspace_id=workspace_id, task_id=task_id),
            start=1,
        ):
            values = detail.field_values if isinstance(detail.field_values, dict) else {}
            expected_parent_documents.append(
                {
                    "raw_record_id": int_value(values.get("raw_record_id")) or int(detail.id),
                    "parent_sequence": parent_sequence,
                }
            )

    coverage = analyze_waybill_coverage(
        expected_parent_documents=expected_parent_documents,
        rows=rows,
        normal_export_count=sum(recognition_report_row_is_exportable(row) for row in rows),
        exception_export_count=len(recognition_exception_export_rows(rows)),
    )
    if not coverage["ok"]:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="采集面单覆盖不完整，已停止生成报货文件。请先处理缺失面单。",
        )
    return coverage


def task_or_404(db: Session, task_id: int, workspace_id: int) -> CaptureTask:
    task = db.scalars(
        select(CaptureTask).where(
            CaptureTask.id == task_id,
            CaptureTask.workspace_id == workspace_id,
            CaptureTask.is_deleted.is_(False),
        )
    ).first()
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Capture task not found.")
    return task


def set_capture_task_archive_state(
    db: Session,
    *,
    task: CaptureTask,
    user_id: int | None,
    archived: bool,
) -> dict[str, int]:
    archived_at = utc_now() if archived else None
    archived_by = user_id if archived else None

    task.archived_at = archived_at
    task.archived_by = archived_by
    task.updated_by = user_id

    raw_records = db.scalars(
        select(RawCaptureRecord).where(
            RawCaptureRecord.workspace_id == task.workspace_id,
            RawCaptureRecord.task_id == task.id,
            RawCaptureRecord.is_deleted.is_(False),
        )
    ).all()
    for record in raw_records:
        record.archived_at = archived_at
        record.archived_by = archived_by
        record.updated_by = user_id

    details = standard_details_for_task(
        db,
        workspace_id=task.workspace_id,
        task_id=task.id,
        include_archived=True,
    )
    for detail in details:
        detail.archived_at = archived_at
        detail.archived_by = archived_by
        detail.updated_by = user_id

    return {
        "raw_record_count": len(raw_records),
        "standard_detail_count": len(details),
    }


def maintenance_cutoff(days_before: int | None) -> datetime | None:
    if days_before is None:
        return None
    return datetime.now(timezone.utc) - timedelta(days=days_before)


def capture_task_time(task: CaptureTask) -> datetime | None:
    parsed_time = parse_utc_datetime(task.ended_at) or parse_utc_datetime(task.started_at)
    if parsed_time is not None:
        return parsed_time
    if task.created_at is None:
        return None
    if task.created_at.tzinfo is None:
        return task.created_at.replace(tzinfo=timezone.utc)
    return task.created_at.astimezone(timezone.utc)


def capture_task_before_cutoff(task: CaptureTask, cutoff: datetime | None) -> bool:
    if cutoff is None:
        return True
    task_time = capture_task_time(task)
    if task_time is None:
        return False
    if task_time.tzinfo is None:
        task_time = task_time.replace(tzinfo=timezone.utc)
    return task_time <= cutoff


def standard_detail_task_id(detail: StandardDetail) -> int | None:
    values = detail.field_values if isinstance(detail.field_values, dict) else {}
    return int_value(values.get("capture_task_id"))


def capture_data_summary(db: Session, *, workspace_id: int) -> dict[str, Any]:
    tasks = db.scalars(
        select(CaptureTask).where(
            CaptureTask.workspace_id == workspace_id,
            CaptureTask.is_deleted.is_(False),
        )
    ).all()
    raw_records = db.scalars(
        select(RawCaptureRecord).where(
            RawCaptureRecord.workspace_id == workspace_id,
            RawCaptureRecord.is_deleted.is_(False),
        )
    ).all()
    details = db.scalars(
        select(StandardDetail).where(
            StandardDetail.workspace_id == workspace_id,
            StandardDetail.is_deleted.is_(False),
        )
    ).all()
    active_tasks = [task for task in tasks if not task.archived_at]
    archived_tasks = [task for task in tasks if task.archived_at]
    archive_ready_tasks = [
        task
        for task in active_tasks
        if task.status != "collecting"
    ]
    return {
        "active": {
            "capture_tasks": len(active_tasks),
            "archive_ready_tasks": len(archive_ready_tasks),
            "raw_records": len([record for record in raw_records if not record.archived_at]),
            "standard_details": len([detail for detail in details if not detail.archived_at]),
        },
        "archived": {
            "capture_tasks": len(archived_tasks),
            "raw_records": len([record for record in raw_records if record.archived_at]),
            "standard_details": len([detail for detail in details if detail.archived_at]),
        },
        "collecting_tasks": len([task for task in active_tasks if task.status == "collecting"]),
    }


def business_download_timestamp(now: datetime | None = None) -> str:
    source_time = now or datetime.now(timezone.utc)
    if source_time.tzinfo is None:
        source_time = source_time.replace(tzinfo=timezone.utc)
    return source_time.astimezone(BUSINESS_DOWNLOAD_TIMEZONE).strftime("%Y%m%d_%H%M%S")


def business_download_filename(
    prefix: str,
    extension: str,
    *,
    timestamp: str | None = None,
) -> str:
    clean_prefix = safe_download_name_part(prefix)
    clean_extension = extension if extension.startswith(".") else f".{extension}"
    return f"{clean_prefix}_{timestamp or business_download_timestamp()}{clean_extension}"


def xlsx_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    quoted_filename = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=download.xlsx; filename*=UTF-8''{quoted_filename}",
        },
    )


def zip_stream_response(buffer: BytesIO, filename: str) -> StreamingResponse:
    buffer.seek(0)
    quoted_filename = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=download.zip; filename*=UTF-8''{quoted_filename}"},
    )


def safe_xlsx_cell_value(value: Any) -> Any:
    if isinstance(value, str) and value.lstrip(" \t\r\n")[:1] in {"=", "+", "-", "@"}:
        return f"'{value}"
    return value


def append_xlsx_rows(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append([safe_xlsx_cell_value(value) for value in headers])
    for row in rows:
        sheet.append([safe_xlsx_cell_value(value) for value in row])
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:80])
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def style_recognition_report_sheet(sheet, layout: dict[str, Any] | None = None) -> None:
    normalized_layout = normalize_recognition_report_layout(layout)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, column in enumerate(visible_recognition_report_columns(normalized_layout), start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = report_layout_width_to_excel_width(column["width"])

    sheet.row_dimensions[1].height = report_layout_height_to_excel_points(normalized_layout["header_row_height"])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif not cell.value:
                cell.value = None

    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = report_layout_height_to_excel_points(normalized_layout["row_height"])


def safe_excel_sheet_title(value: str, used_titles: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?/\\]", "_", text_value(value) or "未设置档口").strip("' ") or "未设置档口"
    title = base[:31]
    suffix = 2
    while title in used_titles:
        marker = f"_{suffix}"
        title = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    used_titles.add(title)
    return title


def safe_download_name_part(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", text_value(value) or "未设置档口").strip() or "未设置档口"


def append_recognition_report_sheet(
    workbook: Workbook,
    *,
    title: str,
    report_rows: list[dict[str, Any]],
    report_layout: dict[str, Any],
    images_by_id: dict[int, ImageAsset],
    image_buffers: list[BytesIO],
    used_titles: set[str],
) -> None:
    sheet = workbook.create_sheet(safe_excel_sheet_title(title, used_titles))
    append_xlsx_rows(
        sheet,
        recognition_report_headers(report_layout),
        recognition_report_export_rows_from_line_items(report_rows, report_layout),
    )
    style_recognition_report_sheet(sheet, report_layout)
    attach_recognition_report_images(sheet, report_rows, images_by_id, image_buffers, report_layout)


def recognition_report_workbook(
    *,
    report_rows: list[dict[str, Any]],
    report_layout: dict[str, Any],
    images_by_id: dict[int, ImageAsset],
    sheet_title: str = "报货表",
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    image_buffers: list[BytesIO] = []
    append_recognition_report_sheet(
        workbook,
        title=sheet_title,
        report_rows=report_rows,
        report_layout=report_layout,
        images_by_id=images_by_id,
        image_buffers=image_buffers,
        used_titles=set(),
    )
    workbook._recognition_image_buffers = image_buffers  # type: ignore[attr-defined]
    return workbook


def append_recognition_exception_sheet(workbook: Workbook, exception_rows: list[list[Any]]):
    sheet = workbook.create_sheet(RECOGNITION_EXCEPTION_SHEET_TITLE)
    append_xlsx_rows(sheet, RECOGNITION_EXCEPTION_HEADERS, exception_rows)
    return sheet


def standard_details_for_task(
    db: Session,
    *,
    workspace_id: int,
    task_id: int,
    include_archived: bool = False,
) -> list[StandardDetail]:
    statement = select(StandardDetail).where(
        StandardDetail.workspace_id == workspace_id,
        StandardDetail.is_deleted.is_(False),
    )
    if not include_archived:
        statement = statement.where(StandardDetail.archived_at.is_(None))
    return [
        detail
        for detail in db.scalars(statement.order_by(StandardDetail.id.asc())).all()
        if int((detail.field_values or {}).get("capture_task_id") or 0) == task_id
    ]


def get_collector_from_token(
    db: Session,
    x_collector_token: str | None,
) -> Collector:
    if not x_collector_token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing collector token.")

    token_hash = hash_collector_token(x_collector_token)
    collector = db.scalars(
        select(Collector).where(
            Collector.token_hash == token_hash,
            Collector.is_enabled.is_(True),
            Collector.is_deleted.is_(False),
        )
    ).first()
    settings = get_settings()
    previous_key = settings.collector_token_previous_hash_key
    if collector is None and previous_key and previous_key != settings.collector_token_hash_key:
        previous_hash = hash_collector_token(x_collector_token, previous_key)
        collector = db.scalars(
            select(Collector).where(
                Collector.token_hash == previous_hash,
                Collector.is_enabled.is_(True),
                Collector.is_deleted.is_(False),
            )
        ).first()
        if collector is not None:
            collector.token_hash = token_hash
            db.flush()
    if collector is None or (
        isinstance(collector.status_payload, dict)
        and collector.status_payload.get("runtime_status") == "enrollment_pending"
    ):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid collector token.")
    return collector


def lock_collector_runtime_request(
    db: Session,
    x_collector_token: str | None,
    *,
    reject_legacy_during_v2: bool,
) -> Collector:
    authenticated = get_collector_from_token(db, x_collector_token)
    collector_id = authenticated.id
    db.commit()

    statement = update(Collector).where(
        Collector.id == collector_id,
        Collector.is_enabled.is_(True),
        Collector.is_deleted.is_(False),
    )
    if reject_legacy_during_v2:
        now = utc_now()
        statement = statement.where(
            or_(
                Collector.assignment_protocol_version < COLLECTOR_TASK_WINDOW_PROTOCOL,
                Collector.assignment_protocol_lease_expires_at.is_(None),
                Collector.assignment_protocol_lease_expires_at < now,
            )
        )
    result = db.execute(
        statement.values(protocol_revision=Collector.protocol_revision + 1)
    )
    if result.rowcount != 1:
        db.rollback()
        current = get_collector_from_token(db, x_collector_token)
        if reject_legacy_during_v2 and collector_v2_protocol_lease_active(current):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="新版采集器已接管，旧版采集器上传已停止。",
            )
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="采集器状态正在更新，请重试。",
        )

    db.expire_all()
    return get_collector_from_token(db, x_collector_token)


def active_task_statement(workspace_id: int, collector_db_id: int | None = None):
    statement = select(CaptureTask).where(
        CaptureTask.workspace_id == workspace_id,
        CaptureTask.status == "collecting",
        CaptureTask.is_deleted.is_(False),
    )
    if collector_db_id is not None:
        statement = statement.where(
            (CaptureTask.collector_id.is_(None)) | (CaptureTask.collector_id == collector_db_id)
        )
    return statement.order_by(CaptureTask.id.desc())


def collector_task_windows(
    db: Session,
    *,
    workspace_id: int,
    collector_db_id: int,
    pending_captured_at: str | None,
    pending_captured_until: str | None,
    pending_row_count: int,
) -> tuple[list[CaptureTask], bool]:
    active = db.scalars(active_task_statement(workspace_id, collector_db_id)).all()
    if pending_row_count <= 0:
        return active, True

    statement = select(CaptureTask).where(
        CaptureTask.workspace_id == workspace_id,
        CaptureTask.status == "completed",
        CaptureTask.is_deleted.is_(False),
        (CaptureTask.collector_id.is_(None)) | (CaptureTask.collector_id == collector_db_id),
    )
    pending_from = parse_collector_datetime(pending_captured_at)
    pending_until = parse_collector_datetime(pending_captured_until)
    coverage_complete = pending_from is not None and pending_until is not None
    if coverage_complete:
        statement = statement.where(
            CaptureTask.started_at <= (pending_until + timedelta(seconds=1)).isoformat(),
            CaptureTask.ended_at >= pending_from.isoformat(),
        ).order_by(
            CaptureTask.id.asc(),
        )
    else:
        statement = statement.order_by(CaptureTask.id.desc()).limit(1)
    completed = db.scalars(statement).all()
    by_id = {task.id: task for task in [*completed, *active]}
    return [by_id[task_id] for task_id in sorted(by_id)], coverage_complete


def collector_has_task_window_lease(collector: Collector, task_id: int) -> bool:
    payload = collector.status_payload if isinstance(collector.status_payload, dict) else {}
    lease = payload.get("task_window_lease")
    if not isinstance(lease, dict) or task_id not in set(lease.get("task_ids") or []):
        return False
    expires_at = parse_utc_datetime(str(lease.get("expires_at") or ""))
    return expires_at is not None and expires_at >= datetime.now(timezone.utc)


def collector_v2_protocol_lease_active(collector: Collector) -> bool:
    if collector.assignment_protocol_version < COLLECTOR_TASK_WINDOW_PROTOCOL:
        return False
    expires_at = parse_utc_datetime(collector.assignment_protocol_lease_expires_at)
    return expires_at is not None and expires_at >= datetime.now(timezone.utc)


def start_collector_enrollment(collector: Collector) -> str:
    token = create_collector_token()
    status_payload = collector.status_payload if isinstance(collector.status_payload, dict) else {}
    collector.token_hash = hash_collector_token(token)
    collector.online_status = "offline"
    collector.status_payload = {
        **status_payload,
        "runtime_status": "enrollment_pending",
        "enrollment_expires_at": (
            datetime.now(timezone.utc) + COLLECTOR_ENROLLMENT_TIMEOUT
        ).isoformat(),
    }
    return token


def upsert_collector(
    db: Session,
    *,
    tenant_id: int | None,
    workspace_id: int,
    payload: CollectorRegisterRequest,
    user_id: int | None,
) -> tuple[Collector, str]:
    identity_token = create_collector_token()
    collector_identity = clean_optional_text(payload.collector_id) or f"collector-{identity_token[:12]}"
    source_machine = clean_optional_text(payload.source_machine) or None
    display_name = collector_display_name(
        payload.collector_name,
        source_machine=source_machine,
        collector_id=collector_identity,
    )

    collector = db.scalars(
        select(Collector).where(
            Collector.workspace_id == workspace_id,
            Collector.collector_id == collector_identity,
            Collector.is_deleted.is_(False),
        )
    ).first()
    if collector is None:
        collector = Collector(
            tenant_id=tenant_id,
            workspace_id=workspace_id,
            collector_id=collector_identity,
            collector_name=display_name,
            source_machine=source_machine,
            client_version=payload.client_version,
            online_status="offline",
            remark=payload.remark,
            created_by=user_id,
            updated_by=user_id,
        )
        db.add(collector)
    else:
        if is_default_collector_display_name(collector.collector_name) or not is_default_collector_display_name(
            payload.collector_name
        ):
            collector.collector_name = display_name
        collector.source_machine = source_machine
        collector.client_version = payload.client_version
        collector.is_enabled = True
        collector.remark = payload.remark
        collector.updated_by = user_id

    token = start_collector_enrollment(collector)
    db.commit()
    db.refresh(collector)
    return collector, token


def collector_identity_is_available(
    db: Session,
    *,
    workspace_id: int,
    collector_identity: str,
    current_collector_id: int,
) -> bool:
    existing = db.scalars(
        select(Collector).where(
            Collector.workspace_id == workspace_id,
            Collector.collector_id == collector_identity,
            Collector.id != current_collector_id,
            Collector.is_deleted.is_(False),
        )
    ).first()
    return existing is None


@router.get("/collector-client/download")
def download_collector_client(
    mode: str = Query(default="cli", pattern="^(cli|script|exe)$"),
    _current_user: CurrentUser = Depends(get_current_user),
) -> Response:
    if mode == "exe":
        exe_path, _, _ = require_collector_client_release(collector_client_source_dir())
        content = exe_path.read_bytes()
        filename = quote("Cargo Platform 采集器.exe")
        return Response(
            content=content,
            media_type="application/vnd.microsoft.portable-executable",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
                "Content-Length": str(len(content)),
            },
        )

    archive = build_collector_client_archive(mode)
    content = archive.getvalue()
    filename = quote("订单整理系统采集器.zip")
    return Response(
        content=content,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "Content-Length": str(len(content)),
        },
    )


@router.post("/collector-control/register", status_code=status.HTTP_201_CREATED)
def register_collector(
    payload: CollectorRegisterRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    tenant_id = get_workspace_tenant_id(db, workspace_id)
    collector, token = upsert_collector(
        db,
        tenant_id=tenant_id,
        workspace_id=workspace_id,
        payload=payload,
        user_id=current_user.id,
    )
    return {
        "collector": public_collector(collector),
        "connection_code": build_connection_code(payload.public_base_url, token),
    }


@router.post("/collector-runtime/enroll")
def enroll_collector(
    payload: CollectorEnrollmentRequest,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    pending_hash = hash_collector_token(payload.token)
    collector = db.scalars(
        select(Collector).where(
            Collector.token_hash == pending_hash,
            Collector.is_enabled.is_(True),
            Collector.is_deleted.is_(False),
        )
    ).first()
    status_payload = (
        dict(collector.status_payload)
        if collector is not None and isinstance(collector.status_payload, dict)
        else {}
    )
    expires_at = parse_utc_datetime(str(status_payload.get("enrollment_expires_at") or ""))
    if (
        collector is None
        or status_payload.get("runtime_status") != "enrollment_pending"
        or expires_at is None
    ):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid enrollment token.")
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=status.HTTP_410_GONE, detail="Enrollment token expired.")

    collector_identity = clean_optional_text(payload.collector_id) or collector.collector_id
    if not collector_identity_is_available(
        db,
        workspace_id=collector.workspace_id,
        collector_identity=collector_identity,
        current_collector_id=collector.id,
    ):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Collector identity is already in use.")

    source_machine = clean_optional_text(payload.source_machine) or collector.source_machine
    collector_name = collector.collector_name
    reported_name = collector_display_name(
        payload.collector_name,
        source_machine=source_machine,
        collector_id=collector_identity,
    )
    if is_default_collector_display_name(collector_name) or not is_default_collector_display_name(
        payload.collector_name
    ):
        collector_name = reported_name

    status_payload.pop("enrollment_expires_at", None)
    status_payload["runtime_status"] = "enrolled"
    status_payload["enrolled_at"] = utc_now()
    device_token = create_collector_token()
    result = db.execute(
        update(Collector)
        .where(
            Collector.id == collector.id,
            Collector.token_hash == pending_hash,
            Collector.is_enabled.is_(True),
            Collector.is_deleted.is_(False),
        )
        .values(
            token_hash=hash_collector_token(device_token),
            collector_id=collector_identity,
            collector_name=collector_name,
            source_machine=source_machine,
            client_version=payload.client_version or collector.client_version,
            status_payload=status_payload,
        )
    )
    if result.rowcount != 1:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid enrollment token.")
    collector_id = collector.id
    db.commit()
    db.expire_all()
    enrolled = db.get(Collector, collector_id)
    if enrolled is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid enrollment token.")
    return {"collector": public_collector(enrolled), "collector_token": device_token}


@router.post("/collector-control/{collector_id}/repair-code")
def repair_collector_connection_code(
    collector_id: int,
    payload: CollectorRepairCodeRequest,
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    collector = db.get(Collector, collector_id)
    if collector is None or collector.workspace_id != workspace_id or collector.is_deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collector not found.")
    token = start_collector_enrollment(collector)
    collector.is_enabled = True
    collector.updated_by = current_user.id
    db.commit()
    db.refresh(collector)
    return {
        "collector": public_collector(collector),
        "connection_code": build_connection_code(payload.public_base_url, token),
    }


@router.get("/collector-control/status")
def collector_status(
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    collectors = db.scalars(
        select(Collector)
        .where(Collector.workspace_id == workspace_id, Collector.is_deleted.is_(False))
        .order_by(Collector.id.desc())
    ).all()
    active_task = db.scalars(active_task_statement(workspace_id)).first()
    return {
        "collectors": [public_collector(collector) for collector in collectors],
        "active_task": public_task(active_task) if active_task else None,
        "collector_client": collector_client_release_status(),
    }


@router.post("/collector-control/start", status_code=status.HTTP_201_CREATED)
def start_capture(
    payload: CaptureStartRequest = Body(default_factory=CaptureStartRequest),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    existing = db.scalars(active_task_statement(workspace_id)).first()
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A capture task is already collecting.")

    tenant_id = get_workspace_tenant_id(db, workspace_id)
    if payload.collector_id is not None:
        collector = db.get(Collector, payload.collector_id)
        if collector is None or collector.workspace_id != workspace_id or collector.is_deleted:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Collector access denied.")

    task = CaptureTask(
        tenant_id=tenant_id,
        workspace_id=workspace_id,
        name=payload.name or f"采集任务 {utc_now()}",
        collector_id=payload.collector_id,
        status="collecting",
        started_at=utc_now(),
        config={"started_by": current_user.id},
        created_by=current_user.id,
        updated_by=current_user.id,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return public_task(task)


@router.post("/collector-control/stop")
def stop_capture(
    payload: CaptureStopRequest = Body(default_factory=CaptureStopRequest),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    if payload.task_id is None:
        task = db.scalars(active_task_statement(workspace_id)).first()
    else:
        task = db.scalars(
            select(CaptureTask).where(
                CaptureTask.id == payload.task_id,
                CaptureTask.workspace_id == workspace_id,
                CaptureTask.is_deleted.is_(False),
            )
        ).first()
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Active capture task not found.")
    if task.status != "collecting":
        return public_task(task)

    config = dict(task.config or {})
    config["ended_by"] = current_user.id
    task.config = config
    task.status = "completed"
    task.ended_at = utc_now()
    task.updated_by = current_user.id
    db.commit()
    db.refresh(task)
    return public_task(task)


@router.get("/system-settings/data-maintenance")
def data_maintenance_summary(
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    return capture_data_summary(db, workspace_id=workspace_id)


@router.post("/system-settings/data-maintenance/archive-capture-data")
def archive_capture_data(
    payload: ArchiveCaptureDataRequest = Body(default_factory=ArchiveCaptureDataRequest),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    cutoff = maintenance_cutoff(payload.days_before)
    tasks = [
        task
        for task in db.scalars(
            select(CaptureTask)
            .where(
                CaptureTask.workspace_id == workspace_id,
                CaptureTask.is_deleted.is_(False),
                CaptureTask.archived_at.is_(None),
                CaptureTask.status != "collecting",
            )
            .order_by(CaptureTask.id.asc())
        ).all()
        if capture_task_before_cutoff(task, cutoff)
    ]

    archived_raw_records = 0
    archived_standard_details = 0
    for task in tasks:
        counts = set_capture_task_archive_state(
            db,
            task=task,
            user_id=current_user.id,
            archived=True,
        )
        archived_raw_records += counts["raw_record_count"]
        archived_standard_details += counts["standard_detail_count"]

    db.commit()
    return {
        "archived_capture_tasks": len(tasks),
        "archived_raw_records": archived_raw_records,
        "archived_standard_details": archived_standard_details,
        "summary": capture_data_summary(db, workspace_id=workspace_id),
    }


@router.post("/system-settings/data-maintenance/delete-archived-capture-data")
def delete_archived_capture_data(
    payload: DeleteArchivedCaptureDataRequest,
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    if payload.confirm_text.strip() != "删除归档数据":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="请输入确认文字：删除归档数据",
        )

    cutoff = maintenance_cutoff(payload.days_before)
    tasks = [
        task
        for task in db.scalars(
            select(CaptureTask)
            .where(
                CaptureTask.workspace_id == workspace_id,
                CaptureTask.is_deleted.is_(False),
                CaptureTask.archived_at.is_not(None),
            )
            .order_by(CaptureTask.id.asc())
        ).all()
        if capture_task_before_cutoff(task, cutoff)
    ]
    task_ids = {int(task.id) for task in tasks}
    raw_records = db.scalars(
        select(RawCaptureRecord).where(
            RawCaptureRecord.workspace_id == workspace_id,
            RawCaptureRecord.task_id.in_(task_ids),
            RawCaptureRecord.archived_at.is_not(None),
        )
    ).all() if task_ids else []
    details = [
        detail
        for detail in db.scalars(
            select(StandardDetail).where(
                StandardDetail.workspace_id == workspace_id,
                StandardDetail.archived_at.is_not(None),
            )
        ).all()
        if (standard_detail_task_id(detail) or 0) in task_ids
    ] if task_ids else []
    detail_batch_ids = {
        int(detail.standard_detail_batch_id)
        for detail in details
        if detail.standard_detail_batch_id
    }
    detail_batches = db.scalars(
        select(StandardDetailBatch).where(
            StandardDetailBatch.workspace_id == workspace_id,
            StandardDetailBatch.id.in_(detail_batch_ids),
        )
    ).all() if detail_batch_ids else []

    deleted_counts = {
        "deleted_capture_tasks": len(tasks),
        "deleted_raw_records": len(raw_records),
        "deleted_standard_details": len(details),
        "deleted_standard_detail_batches": len(detail_batches),
    }
    for detail in details:
        db.delete(detail)
    for batch in detail_batches:
        db.delete(batch)
    for record in raw_records:
        db.delete(record)
    for task in tasks:
        db.delete(task)

    db.commit()
    return {
        **deleted_counts,
        "summary": capture_data_summary(db, workspace_id=workspace_id),
    }


@router.post("/collector-control/parse-records")
def parse_capture_records(
    payload: ParseRecordsRequest = Body(default_factory=ParseRecordsRequest),
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    statement = select(RawCaptureRecord).where(
        RawCaptureRecord.workspace_id == workspace_id,
        RawCaptureRecord.is_deleted.is_(False),
        RawCaptureRecord.archived_at.is_(None),
    )
    if payload.task_id is not None:
        statement = statement.where(RawCaptureRecord.task_id == payload.task_id)
    records = db.scalars(statement.order_by(RawCaptureRecord.id.asc())).all()

    active_pack = active_recognition_rule_pack(db, workspace_id=workspace_id)
    if active_pack is None:
        return {
            "status": RULE_PACK_MISSING_STATUS,
            "rule_pack_required": True,
            "message": "当前工作空间未启用识别规则包。请先导入并启用规则包，再进行面单识别。",
            "parsed": 0,
            "skipped": 0,
            "raw_record_count": len(records),
            "task_id": payload.task_id,
        }

    raise HTTPException(
        status_code=status.HTTP_410_GONE,
        detail="旧面单解析入口已停用。请使用独立面单解析服务生成订单行，避免旧 standard_details 与新订单行数据混用。",
    )


@router.get("/collector-control/tasks/{task_id}/raw-document")
def download_raw_capture_document(
    task_id: int,
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> StreamingResponse:
    task = task_or_404(db, task_id, workspace_id)
    records = db.scalars(
        select(RawCaptureRecord)
        .where(
            RawCaptureRecord.workspace_id == workspace_id,
            RawCaptureRecord.task_id == task.id,
            RawCaptureRecord.is_deleted.is_(False),
        )
        .order_by(RawCaptureRecord.id.asc())
    ).all()
    collector_ids = sorted({int(record.collector_id) for record in records if record.collector_id})
    collectors_by_id = {
        collector.id: collector
        for collector in db.scalars(
            select(Collector).where(
                Collector.id.in_(collector_ids),
                Collector.workspace_id == workspace_id,
                Collector.is_deleted.is_(False),
            )
        ).all()
    } if collector_ids else {}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原文"
    rows = [
        [
            record.id,
            raw_record_collector_label(record, collectors_by_id),
            record.source_machine,
            source_component_label(record.source_component),
            record.source_index,
            record.dedupe_key,
            record.captured_at,
            record.payload_format,
            json_text(record.source_columns),
            record.status,
            record.raw_payload,
        ]
        for record in records
    ]
    append_xlsx_rows(
        sheet,
        [
            "ID",
            "采集器",
            "电脑名",
            "来源组件",
            "来源序号",
            "去重键",
            "采集时间",
            "原文格式",
            "本地来源信息",
            "状态",
            "采集原文",
        ],
        rows,
    )
    return xlsx_response(workbook, business_download_filename("采集原文", "xlsx"))


@router.get("/collector-control/tasks/{task_id}/standard-document")
def download_standard_capture_document(
    task_id: int,
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> StreamingResponse:
    task = task_or_404(db, task_id, workspace_id)
    details = standard_details_for_task(db, workspace_id=workspace_id, task_id=task.id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "整理结果"
    export_fields = db.scalars(
        select(ExportHeaderDefinition)
        .where(
            ExportHeaderDefinition.workspace_id == workspace_id,
            ExportHeaderDefinition.export_enabled.is_(True),
            ExportHeaderDefinition.export_order > 0,
            ExportHeaderDefinition.is_deleted.is_(False),
        )
        .order_by(ExportHeaderDefinition.export_order.asc(), ExportHeaderDefinition.id.asc())
    ).all()

    if not export_fields:
        append_xlsx_rows(
            sheet,
            ["提示"],
            [["当前工作区还没有定义整理文档表头，暂不生成业务整理文档。"]],
        )
        return xlsx_response(workbook, business_download_filename("整理文档", "xlsx"))

    rows = []
    for detail in details:
        for values in standard_detail_export_rows(detail):
            rows.append(
                [export_field_value(field.code, values) for field in export_fields]
            )
    append_xlsx_rows(
        sheet,
        [field.name for field in export_fields],
        rows,
    )
    return xlsx_response(workbook, business_download_filename("整理文档", "xlsx"))


@router.get("/collector-control/tasks/{task_id}/report-preview")
@router.get("/collector-control/tasks/{task_id}/recognition-preview")
def preview_capture_task_recognition(
    task_id: int,
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    task = task_or_404(db, task_id, workspace_id)
    details = standard_details_for_task(db, workspace_id=workspace_id, task_id=task.id)
    rows = recognition_rows_for_task(db, workspace_id=workspace_id, task_id=task.id)
    coverage = require_complete_recognition_coverage(
        db,
        workspace_id=workspace_id,
        task_id=task.id,
        rows=rows,
    )
    collected_waybill_count = recognition_expected_waybill_count(
        db,
        workspace_id=workspace_id,
        task_id=task.id,
    )
    covered_waybill_count = recognition_waybill_count(rows)
    summary = export_recognition_summary(rows)
    return {
        "task_id": task.id,
        "task_name": task.name,
        "contract": EXPORT_PRODUCT_SKU_LINKING_CONTRACT,
        "data_source": "order_row_drafts",
        "detail_count": collected_waybill_count or len(details),
        "waybill_count": collected_waybill_count,
        "collected_waybill_count": collected_waybill_count,
        "covered_waybill_count": covered_waybill_count,
        "coverage_complete": bool(coverage["ok"]),
        "coverage": coverage,
        "order_row_count": summary["total"],
        "rows": rows,
        "summary": summary,
    }


@router.get("/collector-control/tasks/{task_id}/report-workbook")
@router.get("/collector-control/tasks/{task_id}/recognition-report")
def download_capture_task_recognition_report(
    task_id: int,
    layout: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> StreamingResponse:
    task = task_or_404(db, task_id, workspace_id)
    rows = recognition_rows_for_task(db, workspace_id=workspace_id, task_id=task.id)
    require_complete_recognition_coverage(
        db,
        workspace_id=workspace_id,
        task_id=task.id,
        rows=rows,
    )
    images_by_id = recognition_report_image_assets(db, workspace_id=workspace_id, rows=rows)
    report_layout = recognition_report_layout_from_query(layout)
    report_rows = recognition_report_line_items(rows, report_layout)
    exception_rows = recognition_exception_export_rows(rows)
    download_timestamp = business_download_timestamp()

    if report_layout["output_mode"] == "stall_workbooks":
        archive = BytesIO()
        with ZipFile(archive, "w", ZIP_DEFLATED) as zip_file:
            for stall_name, stall_rows in recognition_report_rows_by_stall(report_rows).items():
                stall_workbook = recognition_report_workbook(
                    report_rows=stall_rows,
                    report_layout=report_layout,
                    images_by_id=images_by_id,
                    sheet_title=safe_download_name_part(stall_name),
                )
                workbook_buffer = BytesIO()
                stall_workbook.save(workbook_buffer)
                zip_file.writestr(
                    business_download_filename(
                        f"{safe_download_name_part(stall_name)}_{BUSINESS_REPORT_DOWNLOAD_PREFIX}",
                        "xlsx",
                        timestamp=download_timestamp,
                    ),
                    workbook_buffer.getvalue(),
                )
            exception_workbook = Workbook()
            exception_sheet = exception_workbook.active
            exception_sheet.title = RECOGNITION_EXCEPTION_SHEET_TITLE
            append_xlsx_rows(exception_sheet, RECOGNITION_EXCEPTION_HEADERS, exception_rows)
            exception_buffer = BytesIO()
            exception_workbook.save(exception_buffer)
            zip_file.writestr(
                business_download_filename(RECOGNITION_EXCEPTION_SHEET_TITLE, "xlsx", timestamp=download_timestamp),
                exception_buffer.getvalue(),
            )
        return zip_stream_response(
            archive,
            business_download_filename(f"{BUSINESS_REPORT_DOWNLOAD_PREFIX}_分档口", "zip", timestamp=download_timestamp),
        )

    workbook = Workbook()
    image_buffers: list[BytesIO] = []
    if report_layout["output_mode"] == "stall_sheet":
        workbook.remove(workbook.active)
        used_titles: set[str] = set()
        for stall_name, stall_rows in recognition_report_rows_by_stall(report_rows).items():
            append_recognition_report_sheet(
                workbook,
                title=stall_name,
                report_rows=stall_rows,
                report_layout=report_layout,
                images_by_id=images_by_id,
                image_buffers=image_buffers,
                used_titles=used_titles,
            )
    else:
        sheet = workbook.active
        sheet.title = "报货表"
        append_xlsx_rows(
            sheet,
            recognition_report_headers(report_layout),
            recognition_report_export_rows_from_line_items(report_rows, report_layout),
        )
        style_recognition_report_sheet(sheet, report_layout)
        attach_recognition_report_images(sheet, report_rows, images_by_id, image_buffers, report_layout)

    append_recognition_exception_sheet(workbook, exception_rows)
    return xlsx_response(
        workbook,
        business_download_filename(BUSINESS_REPORT_DOWNLOAD_PREFIX, "xlsx", timestamp=download_timestamp),
    )


@router.post("/collector-runtime/heartbeat")
def collector_heartbeat(
    payload: CollectorHeartbeatRequest = Body(default_factory=CollectorHeartbeatRequest),
    db: Session = Depends(get_db),
    x_collector_token: Annotated[str | None, Header(alias="X-Collector-Token")] = None,
) -> dict[str, Any]:
    collector = lock_collector_runtime_request(
        db,
        x_collector_token,
        reject_legacy_during_v2=False,
    )
    received_datetime = datetime.now(timezone.utc)
    received_at = received_datetime.isoformat()
    collector.online_status = "online"
    collector.last_heartbeat_at = received_at
    previous_status_payload = (
        collector.status_payload if isinstance(collector.status_payload, dict) else {}
    )
    status_payload = {
        "runtime_status": payload.runtime_status or "unknown",
        "adapter_status": payload.adapter_status or {},
        "queue_size": payload.queue_size,
        "last_error": payload.last_error,
        "last_upload_at": payload.last_upload_at,
        "last_reconnect_reason": payload.last_reconnect_reason,
        "received_at": received_at,
    }
    if payload.source_machine:
        collector.source_machine = payload.source_machine
    reported_identity = str(payload.collector_id or payload.source_machine or "").strip()
    if reported_identity and collector_identity_is_available(
        db,
        workspace_id=collector.workspace_id,
        collector_identity=reported_identity,
        current_collector_id=collector.id,
    ):
        collector.collector_id = reported_identity
    reported_display_name = collector_display_name(
        payload.collector_name,
        source_machine=payload.source_machine,
        collector_id=reported_identity or collector.collector_id,
    )
    if is_default_collector_display_name(collector.collector_name) or not is_default_collector_display_name(
        payload.collector_name
    ):
        collector.collector_name = reported_display_name
    if payload.client_version:
        collector.client_version = payload.client_version

    tasks = db.scalars(active_task_statement(collector.workspace_id, collector.id)).all()
    if payload.assignment_protocol_version < COLLECTOR_TASK_WINDOW_PROTOCOL:
        tracked_task_ids = {task_id for task_id in payload.tracked_task_ids if task_id > 0}
        if tracked_task_ids:
            completed_tasks = db.scalars(
                select(CaptureTask).where(
                    CaptureTask.workspace_id == collector.workspace_id,
                    CaptureTask.id.in_(tracked_task_ids),
                    CaptureTask.status == "completed",
                    CaptureTask.archived_at.is_(None),
                    CaptureTask.is_deleted.is_(False),
                    (CaptureTask.collector_id.is_(None))
                    | (CaptureTask.collector_id == collector.id),
                )
            ).all()
            active_ids = {task.id for task in tasks}
            tasks.extend(task for task in completed_tasks if task.id not in active_ids)
    task_windows: list[CaptureTask] = []
    window_coverage_complete = False
    if payload.assignment_protocol_version >= COLLECTOR_TASK_WINDOW_PROTOCOL:
        task_windows, window_coverage_complete = collector_task_windows(
            db,
            workspace_id=collector.workspace_id,
            collector_db_id=collector.id,
            pending_captured_at=payload.pending_captured_at,
            pending_captured_until=payload.pending_captured_until,
            pending_row_count=payload.pending_row_count,
        )
        status_payload["task_window_lease"] = {
            "task_ids": [task.id for task in task_windows],
            "expires_at": (received_datetime + COLLECTOR_TASK_WINDOW_LEASE).isoformat(),
        }
        collector.assignment_protocol_version = COLLECTOR_TASK_WINDOW_PROTOCOL
        collector.assignment_protocol_lease_expires_at = (
            received_datetime + COLLECTOR_PROTOCOL_LEASE
        ).isoformat()
        status_payload["assignment_protocol_lease"] = {
            "version": COLLECTOR_TASK_WINDOW_PROTOCOL,
            "expires_at": collector.assignment_protocol_lease_expires_at,
        }
    elif collector_v2_protocol_lease_active(collector):
        status_payload["assignment_protocol_lease"] = {
            "version": collector.assignment_protocol_version,
            "expires_at": collector.assignment_protocol_lease_expires_at,
        }
        if "task_window_lease" in previous_status_payload:
            status_payload["task_window_lease"] = previous_status_payload["task_window_lease"]
    collector.status_payload = status_payload
    db.commit()
    return {
        "collector": public_collector(collector),
        "tasks": [public_task(task) for task in tasks],
        "assignment_protocol_version": min(
            payload.assignment_protocol_version,
            COLLECTOR_TASK_WINDOW_PROTOCOL,
        ),
        "task_windows": [public_task(task) for task in task_windows],
        "window_coverage_complete": window_coverage_complete,
    }


@router.post("/collector-runtime/raw-records", status_code=status.HTTP_201_CREATED)
def upload_raw_records(
    payload: RawCaptureBatchRequest,
    db: Session = Depends(get_db),
    x_collector_token: Annotated[str | None, Header(alias="X-Collector-Token")] = None,
) -> dict[str, int]:
    """Persist collector payloads as raw_capture_record only.

    Waybill reading/parsing is owned by the downstream module and is triggered
    explicitly through /collector-control/parse-records.
    """
    collector = lock_collector_runtime_request(
        db,
        x_collector_token,
        reject_legacy_during_v2=(
            payload.assignment_protocol_version < COLLECTOR_TASK_WINDOW_PROTOCOL
        ),
    )
    if (
        payload.assignment_protocol_version >= COLLECTOR_TASK_WINDOW_PROTOCOL
        and not collector_v2_protocol_lease_active(collector)
    ):
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="新版采集器租约已失效，请重新连接后再上传。",
        )
    task = db.get(CaptureTask, payload.task_id)
    leased_archive = bool(
        task is not None
        and task.archived_at
        and not task.is_deleted
        and payload.assignment_protocol_version >= COLLECTOR_TASK_WINDOW_PROTOCOL
        and collector_has_task_window_lease(collector, task.id)
    )
    if (
        task is None
        or task.workspace_id != collector.workspace_id
        or task.is_deleted
        or (task.archived_at and not leased_archive)
        or task.status not in {"collecting", "completed"}
        or (task.collector_id is not None and task.collector_id != collector.id)
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Capture task access denied.")

    inserted = 0
    duplicates = 0
    window_rejected = 0
    for item in payload.records:
        if not collector_record_is_inside_task_window(item, task):
            window_rejected += 1
            continue
        capture_event_key = None
        if item.source_component and item.source_index:
            # A source row is one print event. Equal payloads may be legitimate reprints.
            duplicate_conditions = [
                RawCaptureRecord.workspace_id == collector.workspace_id,
                RawCaptureRecord.collector_id == collector.id,
                RawCaptureRecord.source_component == item.source_component,
                RawCaptureRecord.source_index == item.source_index,
                RawCaptureRecord.is_deleted.is_(False),
            ]
            if payload.assignment_protocol_version < COLLECTOR_TASK_WINDOW_PROTOCOL:
                duplicate_conditions.append(RawCaptureRecord.task_id == task.id)
            existing = db.scalars(select(RawCaptureRecord).where(*duplicate_conditions)).first()
            if existing is not None:
                duplicates += 1
                continue
            if payload.assignment_protocol_version >= COLLECTOR_TASK_WINDOW_PROTOCOL:
                capture_event_key = hashlib.sha256(
                    (
                        f"{collector.workspace_id}\0{collector.id}\0"
                        f"{item.source_component}\0{item.source_index}"
                    ).encode("utf-8")
                ).hexdigest()

        record = build_raw_capture_record(
            collector=collector,
            task=task,
            payload=item,
            captured_at=utc_now(),
        )
        if task.archived_at:
            record.archived_at = task.archived_at
            record.archived_by = task.archived_by
        record.capture_event_key = capture_event_key
        try:
            with db.begin_nested():
                db.add(record)
                db.flush()
        except IntegrityError:
            existing_event = db.scalar(
                select(RawCaptureRecord.id).where(
                    RawCaptureRecord.workspace_id == collector.workspace_id,
                    RawCaptureRecord.capture_event_key == capture_event_key,
                )
            )
            if capture_event_key is None or existing_event is None:
                raise
            duplicates += 1
            continue
        inserted += 1

    db.commit()
    result = {
        "inserted": inserted,
        "skipped": duplicates + window_rejected,
    }
    if payload.assignment_protocol_version >= COLLECTOR_TASK_WINDOW_PROTOCOL:
        result.update(
            {
                "duplicates": duplicates,
                "window_rejected": window_rejected,
            }
        )
    return result
