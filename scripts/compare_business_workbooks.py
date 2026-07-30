from __future__ import annotations

from argparse import ArgumentParser
from datetime import date, datetime, time
from hashlib import sha256
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BUSINESS_HEADERS = (
    "商品",
    "销售属性1",
    "图片",
    "销售属性2",
    "数量",
    "备注",
    "图片匹配文本",
)


def _cell_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value)


def _sheet_rows(sheet: Any) -> list[list[str | int | float | bool | None]]:
    rows: list[list[str | int | float | bool | None]] = []
    for cells in sheet.iter_rows():
        row = [_cell_value(cell.value) for cell in cells]
        while row and row[-1] is None:
            row.pop()
        rows.append(row)
    while rows and not rows[-1]:
        rows.pop()
    return rows


def _image_anchor(image: Any) -> str:
    anchor = image.anchor
    if isinstance(anchor, str):
        return anchor
    return f"{get_column_letter(anchor._from.col + 1)}{anchor._from.row + 1}"


def _image_hashes(sheet: Any) -> list[dict[str, str]]:
    return [
        {"anchor": _image_anchor(image), "sha256": sha256(image._data()).hexdigest()}
        for image in sheet._images
    ]


def workbook_manifest(path: Path) -> dict[str, Any]:
    """Return ordered sheet names, ordered cell rows, and image SHA-256 values."""
    workbook = load_workbook(path, data_only=True)
    try:
        sheets = [
            {
                "name": sheet.title,
                "rows": _sheet_rows(sheet),
                "images": _image_hashes(sheet),
            }
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()
    return {"sheet_names": [sheet["name"] for sheet in sheets], "sheets": sheets}


def compare_workbooks(expected_path: Path, actual_path: Path) -> dict[str, Any]:
    """Return equivalent plus row/column/image differences without comparing ZIP metadata."""
    expected = workbook_manifest(expected_path)
    actual = workbook_manifest(actual_path)
    differences: list[dict[str, Any]] = []
    if expected["sheet_names"] != actual["sheet_names"]:
        differences.append(
            {
                "kind": "sheet_names",
                "expected": expected["sheet_names"],
                "actual": actual["sheet_names"],
            }
        )

    for index, expected_sheet in enumerate(expected["sheets"]):
        if index >= len(actual["sheets"]):
            break
        actual_sheet = actual["sheets"][index]
        expected_rows = expected_sheet["rows"]
        actual_rows = actual_sheet["rows"]
        if len(expected_rows) != len(actual_rows):
            differences.append(
                {
                    "kind": "row_count",
                    "sheet": expected_sheet["name"],
                    "expected": len(expected_rows),
                    "actual": len(actual_rows),
                }
            )
        for row_index, (expected_row, actual_row) in enumerate(zip(expected_rows, actual_rows), start=1):
            if len(expected_row) != len(actual_row):
                differences.append(
                    {
                        "kind": "column_count",
                        "sheet": expected_sheet["name"],
                        "row": row_index,
                        "expected": len(expected_row),
                        "actual": len(actual_row),
                    }
                )
            for column_index, (expected_value, actual_value) in enumerate(
                zip(expected_row, actual_row), start=1
            ):
                if expected_value != actual_value:
                    differences.append(
                        {
                            "kind": "cell_value",
                            "sheet": expected_sheet["name"],
                            "row": row_index,
                            "column": column_index,
                            "expected": expected_value,
                            "actual": actual_value,
                        }
                    )
        if expected_sheet["images"] != actual_sheet["images"]:
            differences.append(
                {
                    "kind": "image_hashes",
                    "sheet": expected_sheet["name"],
                    "expected": expected_sheet["images"],
                    "actual": actual_sheet["images"],
                }
            )
    return {"equivalent": not differences, "differences": differences}


def main(argv: list[str] | None = None) -> int:
    parser = ArgumentParser(description="Compare two business workbooks semantically.")
    parser.add_argument("expected", type=Path)
    parser.add_argument("actual", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args(argv)

    report = compare_workbooks(args.expected, args.actual)
    encoded = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    if args.output is not None:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(encoded, encoding="utf-8")
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    print(encoded, end="")
    return 0 if report["equivalent"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
